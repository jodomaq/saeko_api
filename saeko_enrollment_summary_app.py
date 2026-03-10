"""
Saeko Enrollment Summary App
Aplicación Tkinter para consultar el total de alumnos inscritos por plantel
en un rango de periodos, generando una matriz Plantel × Periodo con exportación a Excel.
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import requests
from auth import SERVICE_ACCOUNT, USER_EMAIL, get_access_token

BASE_URL = "https://app.saeko.io/api/v1"


# ============================================================
# CAPA DE API
# ============================================================

def api_get(endpoint: str, access_token: str, params: dict = None) -> dict:
    headers = {"Authorization": f"Bearer {access_token}"}
    response = requests.get(f"{BASE_URL}{endpoint}", headers=headers, params=params, timeout=60)
    response.raise_for_status()
    return response.json()


def get_schools(access_token: str) -> list:
    data = api_get("/core/schools", access_token, {"include_fields": "cct"})
    return data.get("schools", [])


def get_terms_by_school(access_token: str, school_id: int) -> list:
    data = api_get(f"/core/schools/{school_id}/terms", access_token)
    return data.get("terms", [])


def get_enrollment_count(access_token: str, school_id: int, term_id: int) -> int:
    """Obtiene el total de enrollments de un plantel en un term usando limit=1 y leyendo meta.total."""
    params = {
        "limit": 1,
        "offset": 0,
        "filters": f"school_id={school_id}",
    }
    data = api_get(f"/core/terms/{term_id}/enrollments", access_token, params)
    return data.get("meta", {}).get("total", len(data.get("enrollments", [])))


# ============================================================
# APLICACIÓN TKINTER
# ============================================================

class SaekoEnrollmentSummaryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Saeko - Matriz de Inscripciones Plantel × Periodo")
        self.root.geometry("1100x650")
        self.root.resizable(True, True)

        self.access_token = None
        self.schools = []
        self._cancel_event = threading.Event()
        self._sorted_term_names = []   # lista ordenada de nombres únicos de periodos
        self._term_info = {}           # term_name → {begins_at, ends_at, is_current} (datos del primer term encontrado)
        self._school_term_map = {}     # school_id → {term_name → term_id}
        self.selected_term_names = []  # nombres de periodos en el rango seleccionado
        self.matrix = {}               # (school_id, term_name) → count

        self._build_ui()
        self._authenticate()

    def _build_ui(self):
        main = ttk.Frame(self.root, padding=15)
        main.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main, text="Matriz de Inscripciones Plantel × Periodo",
                  font=("Segoe UI", 16, "bold")).pack(pady=(0, 15))

        # --- Filtros de rango de periodos ---
        filter_frame = ttk.LabelFrame(main, text="Rango de Periodos", padding=10)
        filter_frame.pack(fill=tk.X, pady=(0, 10))

        row0 = ttk.Frame(filter_frame)
        row0.pack(fill=tk.X, pady=3)
        ttk.Label(row0, text="Periodo Inicio:", width=18, anchor="w").pack(side=tk.LEFT)
        self.term_start_var = tk.StringVar()
        self.term_start_combo = ttk.Combobox(row0, textvariable=self.term_start_var,
                                              state="readonly", width=65)
        self.term_start_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)

        row1 = ttk.Frame(filter_frame)
        row1.pack(fill=tk.X, pady=3)
        ttk.Label(row1, text="Periodo Fin:", width=18, anchor="w").pack(side=tk.LEFT)
        self.term_end_var = tk.StringVar()
        self.term_end_combo = ttk.Combobox(row1, textvariable=self.term_end_var,
                                            state="readonly", width=65)
        self.term_end_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # --- Botones ---
        btn_frame = ttk.Frame(main)
        btn_frame.pack(fill=tk.X, pady=(5, 10))
        self.btn_query = ttk.Button(btn_frame, text="Consultar Inscripciones",
                                     command=self._on_query)
        self.btn_query.pack(side=tk.LEFT, padx=(0, 10))
        self.btn_export = ttk.Button(btn_frame, text="Exportar a Excel",
                                      command=self._on_export, state="disabled")
        self.btn_export.pack(side=tk.LEFT, padx=(0, 10))
        self.btn_cancel = ttk.Button(btn_frame, text="Cancelar",
                                      command=self._on_cancel, state="disabled")
        self.btn_cancel.pack(side=tk.LEFT)

        # --- Barra de progreso ---
        self.progress = ttk.Progressbar(main, mode="determinate")
        self.progress.pack(fill=tk.X, pady=(0, 5))

        self.status_var = tk.StringVar(value="Iniciando...")
        ttk.Label(main, textvariable=self.status_var,
                  font=("Segoe UI", 9)).pack(anchor="w")

        # --- Tabla de resultados (se reconstruye dinámicamente) ---
        self.tree_frame = ttk.Frame(main)
        self.tree_frame.pack(fill=tk.BOTH, expand=True, pady=(5, 0))

        self.tree = None
        self._build_empty_tree()

    def _build_empty_tree(self):
        """Construye un Treeview vacío inicial."""
        if self.tree:
            self.tree.destroy()
        for w in self.tree_frame.winfo_children():
            w.destroy()

        columns = ("no", "cct", "nombre_plantel")
        self.tree = ttk.Treeview(self.tree_frame, columns=columns, show="headings", height=18)
        self.tree.heading("no", text="#")
        self.tree.heading("cct", text="CCT")
        self.tree.heading("nombre_plantel", text="Nombre del Plantel")
        self.tree.column("no", width=40, anchor="center")
        self.tree.column("cct", width=130, anchor="center")
        self.tree.column("nombre_plantel", width=400, anchor="w")

        vsb = ttk.Scrollbar(self.tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(self.tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.tree_frame.grid_rowconfigure(0, weight=1)
        self.tree_frame.grid_columnconfigure(0, weight=1)

    def _build_matrix_tree(self, term_names):
        """Reconstruye el Treeview con columnas dinámicas por cada periodo."""
        for w in self.tree_frame.winfo_children():
            w.destroy()

        col_ids = ["no", "cct", "nombre_plantel"]
        for idx, name in enumerate(term_names):
            col_ids.append(f"term_{idx}")
        col_ids.append("total")

        self.tree = ttk.Treeview(self.tree_frame, columns=col_ids, show="headings", height=18)

        self.tree.heading("no", text="#")
        self.tree.column("no", width=35, anchor="center", minwidth=35)
        self.tree.heading("cct", text="CCT")
        self.tree.column("cct", width=120, anchor="center", minwidth=80)
        self.tree.heading("nombre_plantel", text="Nombre del Plantel")
        self.tree.column("nombre_plantel", width=280, anchor="w", minwidth=150)

        for idx, name in enumerate(term_names):
            col_id = f"term_{idx}"
            self.tree.heading(col_id, text=name)
            self.tree.column(col_id, width=90, anchor="center", minwidth=60)

        self.tree.heading("total", text="TOTAL")
        self.tree.column("total", width=80, anchor="center", minwidth=60)

        vsb = ttk.Scrollbar(self.tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(self.tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.tree_frame.grid_rowconfigure(0, weight=1)
        self.tree_frame.grid_columnconfigure(0, weight=1)

    # --- Autenticación ---
    def _authenticate(self):
        self.status_var.set("Autenticando...")
        self._disable_controls()

        def task():
            try:
                token_data = get_access_token(SERVICE_ACCOUNT, USER_EMAIL)
                self.access_token = token_data["access_token"]
                self.root.after(0, self._load_schools_and_terms)
            except Exception as e:
                msg = str(e)
                self.root.after(0, lambda msg=msg: self._show_error("Error de autenticación", msg))

        threading.Thread(target=task, daemon=True).start()

    def _load_schools_and_terms(self):
        self.status_var.set("Cargando planteles y periodos...")

        def task():
            try:
                self.schools = get_schools(self.access_token)
                self._school_term_map = {}
                self._term_info = {}

                for i, school in enumerate(self.schools):
                    school_id = school["id"]
                    self.root.after(0, lambda i=i: self.status_var.set(
                        f"Cargando periodos: plantel {i + 1}/{len(self.schools)}..."
                    ))
                    try:
                        terms = get_terms_by_school(self.access_token, school_id)
                        school_map = {}
                        for t in terms:
                            term_name = t.get("name", f"Term {t['id']}")
                            school_map[term_name] = t["id"]
                            # Guardar info del periodo (primera vez que se ve)
                            if term_name not in self._term_info:
                                self._term_info[term_name] = {
                                    "begins_at": t.get("begins_at", ""),
                                    "ends_at": t.get("ends_at", ""),
                                    "is_current": t.get("is_current", False),
                                }
                            # Si algún plantel lo marca como actual, marcarlo
                            elif t.get("is_current"):
                                self._term_info[term_name]["is_current"] = True
                        self._school_term_map[school_id] = school_map
                    except Exception:
                        self._school_term_map[school_id] = {}

                self.root.after(0, self._populate_terms)
            except Exception as e:
                msg = str(e)
                self.root.after(0, lambda msg=msg: self._show_error("Error cargando datos", msg))

        threading.Thread(target=task, daemon=True).start()

    def _populate_terms(self):
        # Ordenar nombres de periodos por begins_at ascendente
        sorted_names = sorted(
            self._term_info.keys(),
            key=lambda n: self._term_info[n].get("begins_at", "")
        )
        self._sorted_term_names = sorted_names

        term_values = []
        for name in sorted_names:
            info = self._term_info[name]
            current = " [ACTUAL]" if info.get("is_current") else ""
            term_values.append(
                f"{name} | {info.get('begins_at', '')} → {info.get('ends_at', '')}{current}"
            )
        self.term_start_combo["values"] = term_values
        self.term_end_combo["values"] = term_values

        # Seleccionar el actual como fin, si existe
        for i, name in enumerate(sorted_names):
            if self._term_info[name].get("is_current"):
                self.term_end_combo.current(i)
                start_idx = max(0, i - 2)
                self.term_start_combo.current(start_idx)
                break

        self.status_var.set(
            f"{len(self.schools)} planteles, {len(sorted_names)} periodos cargados. "
            "Selecciona rango de periodos y pulsa 'Consultar Inscripciones'."
        )
        self._enable_controls()

    # --- Consulta ---
    def _on_query(self):
        start_sel = self.term_start_combo.current()
        end_sel = self.term_end_combo.current()
        if start_sel < 0 or end_sel < 0:
            messagebox.showwarning("Atención", "Selecciona periodo de inicio y de fin.")
            return
        if start_sel > end_sel:
            messagebox.showwarning("Atención",
                                   "El periodo de inicio debe ser anterior o igual al periodo de fin.")
            return

        self.selected_term_names = self._sorted_term_names[start_sel:end_sel + 1]
        num_terms = len(self.selected_term_names)
        num_schools = len(self.schools)
        total_queries = num_schools * num_terms

        self._cancel_event.clear()
        self._disable_controls()
        self.btn_cancel.configure(state="normal")
        self.btn_export.configure(state="disabled")
        self.progress["value"] = 0
        self.matrix = {}
        self.root.after(0, lambda: self._build_matrix_tree(self.selected_term_names))
        self.status_var.set(
            f"Consultando {num_schools} planteles × {num_terms} periodos ({total_queries} consultas)..."
        )

        def task():
            try:
                done = 0
                cancelled = False
                for i, school in enumerate(self.schools):
                    if self._cancel_event.is_set():
                        cancelled = True
                        break
                    school_id = school["id"]
                    school_name = school.get("name", f"School {school_id}")
                    school_cct = school.get("cct", "")
                    row_total = 0
                    school_terms = self._school_term_map.get(school_id, {})

                    for j, term_name in enumerate(self.selected_term_names):
                        if self._cancel_event.is_set():
                            cancelled = True
                            break
                        self.root.after(0, lambda done=done, name=school_name, tname=term_name: (
                            self.status_var.set(
                                f"Plantel {name} — Periodo {tname}  ({done}/{total_queries})"
                            ),
                            self.progress.configure(value=(done / max(total_queries, 1)) * 100)
                        ))

                        term_id = school_terms.get(term_name)
                        if term_id:
                            try:
                                count = get_enrollment_count(self.access_token, school_id, term_id)
                            except Exception:
                                count = 0
                        else:
                            count = 0

                        self.matrix[(school_id, term_name)] = count
                        row_total += count
                        done += 1

                    if cancelled:
                        break

                    # Insertar fila en la tabla
                    values = [i + 1, school_cct, school_name]
                    for term_name in self.selected_term_names:
                        values.append(self.matrix.get((school_id, term_name), 0))
                    values.append(row_total)
                    self.root.after(0, lambda v=tuple(values): self.tree.insert("", tk.END, values=v))

                if cancelled:
                    self.root.after(0, lambda done=done: self._query_cancelled(done, total_queries))
                else:
                    self.root.after(0, lambda: self._query_done())
            except Exception as e:
                msg = str(e)
                self.root.after(0, lambda msg=msg: self._show_error("Error en consulta", msg))
            finally:
                self.root.after(0, self._enable_controls)

        threading.Thread(target=task, daemon=True).start()

    def _on_cancel(self):
        self._cancel_event.set()
        self.btn_cancel.configure(state="disabled")
        self.status_var.set("Cancelando...")

    def _query_cancelled(self, done, total):
        self.btn_cancel.configure(state="disabled")
        self.progress["value"] = (done / max(total, 1)) * 100
        self.status_var.set(
            f"Consulta cancelada: {done}/{total} consultas completadas."
        )
        if self.matrix:
            self.btn_export.configure(state="normal")

    def _query_done(self):
        self.btn_cancel.configure(state="disabled")
        self.progress["value"] = 100
        grand_total = sum(self.matrix.values())
        self.status_var.set(
            f"Consulta completada: {len(self.schools)} planteles × "
            f"{len(self.selected_term_names)} periodos. Total general: {grand_total} inscripciones."
        )
        if self.matrix:
            self.btn_export.configure(state="normal")

    # --- Exportar a Excel ---
    def _on_export(self):
        if not self.matrix:
            messagebox.showwarning("Atención", "No hay datos para exportar. Realiza una consulta primero.")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Guardar Excel como...",
            initialfile="matriz_inscripciones.xlsx"
        )
        if not filepath:
            return

        self.status_var.set("Exportando a Excel...")

        def task():
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from openpyxl.utils import get_column_letter

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Matriz Inscripciones"

                # Estilos
                header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                bold_font = Font(name="Calibri", bold=True, size=11)
                center_align = Alignment(horizontal="center")
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

                term_names = self.selected_term_names
                num_terms = len(term_names)
                total_cols = 3 + num_terms + 1  # #, CCT, Plantel, terms..., TOTAL

                # Título (fila 1)
                t_start = term_names[0]
                t_end = term_names[-1]
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
                title_cell = ws.cell(row=1, column=1,
                                     value=f"Matriz de Inscripciones — {t_start} a {t_end}")
                title_cell.font = Font(name="Calibri", bold=True, size=14)
                title_cell.alignment = Alignment(horizontal="center")

                # Encabezados (fila 3)
                fixed_headers = ["#", "CCT", "Nombre del Plantel"]
                for col, h in enumerate(fixed_headers, 1):
                    cell = ws.cell(row=3, column=col, value=h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

                for j, name in enumerate(term_names):
                    col = 4 + j
                    cell = ws.cell(row=3, column=col, value=name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

                total_col = 4 + num_terms
                cell = ws.cell(row=3, column=total_col, value="TOTAL")
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

                # Datos (filas 4+)
                for i, school in enumerate(self.schools):
                    row = 4 + i
                    school_id = school["id"]
                    ws.cell(row=row, column=1, value=i + 1).border = thin_border
                    ws.cell(row=row, column=2, value=school.get("cct", "")).border = thin_border
                    ws.cell(row=row, column=3, value=school.get("name", "")).border = thin_border

                    row_total = 0
                    for j, term_name in enumerate(term_names):
                        count = self.matrix.get((school_id, term_name), 0)
                        row_total += count
                        c = ws.cell(row=row, column=4 + j, value=count)
                        c.border = thin_border
                        c.alignment = center_align

                    c = ws.cell(row=row, column=total_col, value=row_total)
                    c.border = thin_border
                    c.alignment = center_align
                    c.font = bold_font

                # Fila de totales por periodo
                totals_row = 4 + len(self.schools)
                ws.merge_cells(start_row=totals_row, start_column=1,
                               end_row=totals_row, end_column=3)
                lbl = ws.cell(row=totals_row, column=1, value="TOTAL POR PERIODO")
                lbl.font = bold_font
                lbl.alignment = Alignment(horizontal="right")
                lbl.border = thin_border
                for c in range(2, 4):
                    ws.cell(row=totals_row, column=c).border = thin_border

                grand_total = 0
                for j, term_name in enumerate(term_names):
                    col_total = sum(
                        self.matrix.get((s["id"], term_name), 0) for s in self.schools
                    )
                    grand_total += col_total
                    c = ws.cell(row=totals_row, column=4 + j, value=col_total)
                    c.border = thin_border
                    c.alignment = center_align
                    c.font = bold_font

                c = ws.cell(row=totals_row, column=total_col, value=grand_total)
                c.border = thin_border
                c.alignment = center_align
                c.font = Font(name="Calibri", bold=True, size=12)

                # Ancho de columnas
                ws.column_dimensions["A"].width = 6
                ws.column_dimensions["B"].width = 18
                ws.column_dimensions["C"].width = 45
                for j in range(num_terms):
                    letter = get_column_letter(4 + j)
                    ws.column_dimensions[letter].width = 16
                ws.column_dimensions[get_column_letter(total_col)].width = 12

                wb.save(filepath)

                self.root.after(0, lambda: self.status_var.set(f"Excel exportado: {filepath}"))
                self.root.after(0, lambda: messagebox.showinfo(
                    "Listo",
                    f"Excel exportado exitosamente.\n"
                    f"{len(self.schools)} planteles × {num_terms} periodos\n"
                    f"Guardado en: {filepath}"
                ))
            except ImportError:
                self.root.after(0, lambda: messagebox.showerror(
                    "Módulo faltante",
                    "Se requiere 'openpyxl' para exportar a Excel.\n"
                    "Instálalo con: pip install openpyxl"
                ))
            except Exception as e:
                msg = str(e)
                self.root.after(0, lambda msg=msg: self._show_error("Error exportando Excel", msg))

        threading.Thread(target=task, daemon=True).start()

    # --- Helpers UI ---
    def _disable_controls(self):
        self.term_start_combo.configure(state="disabled")
        self.term_end_combo.configure(state="disabled")
        self.btn_query.configure(state="disabled")

    def _enable_controls(self):
        self.term_start_combo.configure(state="readonly")
        self.term_end_combo.configure(state="readonly")
        self.btn_query.configure(state="normal")

    def _show_error(self, title, message):
        self.status_var.set(f"Error: {message[:80]}")
        messagebox.showerror(title, message)
        self._enable_controls()


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    root = tk.Tk()
    app = SaekoEnrollmentSummaryApp(root)
    root.mainloop()
