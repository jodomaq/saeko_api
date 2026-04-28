"""
Genera la plantilla de alumnos para carga en SISEC usando un Excel como fuente.
Fuente  : reporte_alumnos.xlsx (columna MATRICULA)
Consulta: API Saeko por matricula (enrollment_number)
Destino : 2026_code/resultado/carga YYYY-MM-DD HH-MM-SS.csv
"""

from __future__ import annotations

import json
import logging
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from tqdm import tqdm

# Agregar raiz al path para importar auth.py
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from auth import SERVICE_ACCOUNT, USER_EMAIL, get_access_token
from genera_alumnos_sexto import (
    CARRERAS_CRUCE_CSV,
    DESTINO_FIELDNAMES,
    ERROR_FIELDNAMES,
    RESULTADO_DIR,
    WARNING_FIELDNAMES,
    _get,
    _advertencias_nombre,
    cargar_cruce_carreras,
    csv_a_excel_particionado,
    get_student_curp,
    procesar_enrollment,
    validar_curp_estructura,
    write_csv_rows,
)

API_LIMIT = 5
CHECKPOINT_FILE = RESULTADO_DIR / "checkpoint_reporte_saeko.json"
REPORTE_XLSX_DEFAULT = ROOT / "reporte_alumnos.xlsx"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


def _normalizar_header(value: str) -> str:
    value = (value or "").strip().upper()
    value = value.replace("Á", "A").replace("É", "E").replace("Í", "I")
    value = value.replace("Ó", "O").replace("Ú", "U")
    return re.sub(r"[^A-Z0-9]", "", value)


def cargar_matriculas_desde_excel(path_xlsx: Path) -> list[str]:
    if not path_xlsx.exists():
        raise FileNotFoundError(f"No existe el archivo Excel: {path_xlsx}")

    wb = openpyxl.load_workbook(path_xlsx, read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx_matricula = None
    for i, header in enumerate(headers):
        if _normalizar_header(str(header or "")) == "MATRICULA":
            idx_matricula = i
            break

    if idx_matricula is None:
        wb.close()
        raise ValueError("No se encontro la columna MATRICULA en el Excel")

    matriculas: list[str] = []
    vistos: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw = row[idx_matricula] if idx_matricula < len(row) else None
        if raw is None:
            continue

        mat = str(raw).strip()
        if not mat:
            continue

        # Si el Excel trae numeros como float (ej. 123456.0), se normaliza.
        if re.match(r"^\d+\.0$", mat):
            mat = mat.split(".", 1)[0]

        if mat in vistos:
            continue

        vistos.add(mat)
        matriculas.append(mat)

    wb.close()
    return matriculas


def buscar_enrollment_por_matricula(matricula: str, token: str, api_url: str) -> Optional[dict]:
    url = f"{api_url}/api/v1/core/enrollments"
    params = {
        "filters": f"enrollment_number={matricula}",
        "include_fields": "student,program_name,group_name,school_name,group_shift,school_id",
        "limit": API_LIMIT,
    }
    data = _get(url, token, params)
    enrollments = data.get("enrollments", [])
    if not enrollments:
        return None

    # Priorizar coincidencia exacta por student_id cuando venga embebido.
    for enr in enrollments:
        student = enr.get("student") or {}
        if str(student.get("student_id") or "").strip() == matricula:
            return enr

    return enrollments[0]


def get_school_cct(school_id: int, token: str, api_url: str, cache: dict[int, str]) -> str:
    if school_id in cache:
        return cache[school_id]

    url = f"{api_url}/api/v1/core/schools/{school_id}"
    data = _get(url, token, {"include_fields": "cct"})
    school = data.get("school") or data
    cct = (school.get("cct") or "").strip().upper()
    cache[school_id] = cct
    return cct


def _guardar_checkpoint(
    timestamp: str,
    processed_matriculas: set[str],
    carga_path: Path,
    errores_path: Path,
    warnings_path: Path,
    reporte_path: Path,
) -> None:
    RESULTADO_DIR.mkdir(parents=True, exist_ok=True)
    with open(CHECKPOINT_FILE, "w", encoding="utf-8") as f:
        json.dump(
            {
                "timestamp": timestamp,
                "processed_matriculas": list(processed_matriculas),
                "carga_path": str(carga_path),
                "errores_path": str(errores_path),
                "warnings_path": str(warnings_path),
                "reporte_path": str(reporte_path),
            },
            f,
            ensure_ascii=False,
            indent=2,
        )


def _cargar_checkpoint() -> Optional[dict]:
    if CHECKPOINT_FILE.exists():
        with open(CHECKPOINT_FILE, encoding="utf-8") as f:
            return json.load(f)
    return None


def _eliminar_checkpoint() -> None:
    if CHECKPOINT_FILE.exists():
        CHECKPOINT_FILE.unlink()


def _resolver_reporte_path() -> Path:
    if len(sys.argv) > 1:
        return Path(sys.argv[1]).expanduser().resolve()
    return REPORTE_XLSX_DEFAULT


def main() -> None:
    reporte_path = _resolver_reporte_path()

    log.info("Cargando tabla de cruce de carreras...")
    cruce_por_cct, cruce_por_plan = cargar_cruce_carreras(CARRERAS_CRUCE_CSV)
    log.info("Cruce cargado: %d entradas (CCT+plan), %d por plan", len(cruce_por_cct), len(cruce_por_plan))

    checkpoint = _cargar_checkpoint()
    if checkpoint:
        print(
            f"\nAvance guardado encontrado del {checkpoint['timestamp']} "
            f"({len(checkpoint['processed_matriculas'])} matriculas ya procesadas)."
        )
        resp = input("Continuar donde se quedo? [s/n]: ").strip().lower()
        if resp in ("s", "si", "y", "yes"):
            timestamp = checkpoint["timestamp"]
            processed_matriculas: set[str] = set(checkpoint["processed_matriculas"])
            carga_path = Path(checkpoint["carga_path"])
            errores_path = Path(checkpoint["errores_path"])
            warnings_path = Path(checkpoint["warnings_path"])
            if checkpoint.get("reporte_path"):
                reporte_path = Path(checkpoint["reporte_path"])
            log.info("Reanudando desde %d matriculas ya procesadas.", len(processed_matriculas))
        else:
            _eliminar_checkpoint()
            timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
            processed_matriculas = set()
            carga_path = RESULTADO_DIR / f"carga {timestamp}.csv"
            errores_path = RESULTADO_DIR / f"errores {timestamp}.csv"
            warnings_path = RESULTADO_DIR / f"warnings {timestamp}.csv"
    else:
        timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
        processed_matriculas = set()
        carga_path = RESULTADO_DIR / f"carga {timestamp}.csv"
        errores_path = RESULTADO_DIR / f"errores {timestamp}.csv"
        warnings_path = RESULTADO_DIR / f"warnings {timestamp}.csv"

    log.info("Leyendo matriculas desde Excel: %s", reporte_path)
    matriculas = cargar_matriculas_desde_excel(reporte_path)
    pendientes = [m for m in matriculas if m not in processed_matriculas]
    log.info("Matriculas en Excel: %d | pendientes: %d", len(matriculas), len(pendientes))

    log.info("Autenticando con Saeko...")
    token = get_access_token(SERVICE_ACCOUNT, USER_EMAIL)["access_token"]
    api_url = SERVICE_ACCOUNT["api_url"].rstrip("/")

    total_ok = 0
    total_err = 0
    total_warn = 0
    school_cct_cache: dict[int, str] = {}

    bar = tqdm(pendientes, desc="Matriculas", unit="alumno", colour="cyan")
    for matricula in bar:
        bar.set_postfix_str(matricula)

        resultados: list[dict] = []
        errores: list[dict] = []
        warnings: list[dict] = []

        school_name = ""
        nombre = ""
        paterno = ""
        materno = ""
        raw_curp = ""

        try:
            enr = buscar_enrollment_por_matricula(matricula, token, api_url)
            if not enr:
                raise ValueError(f"No se encontro enrollment para matricula {matricula}")

            student = enr.get("student") or {}
            school_name = enr.get("school_name") or ""
            surnames = student.get("surnames") or []
            nombre = student.get("first_name") or ""
            paterno = surnames[0] if len(surnames) > 0 else student.get("surname", "")
            materno = surnames[1] if len(surnames) > 1 else ""

            school_id = enr.get("school_id")
            if not school_id:
                raise ValueError("Enrollment sin school_id")
            school_cct = get_school_cct(int(school_id), token, api_url, school_cct_cache)

            raw_curp = (student.get("curp") or "").strip()
            if not raw_curp:
                sid = student.get("id") or enr.get("student_id")
                if sid:
                    raw_curp = get_student_curp(int(sid), token, api_url)

            suite_ok, suite_err = validar_curp_estructura(raw_curp)
            if suite_ok:
                try:
                    resultados.append(
                        procesar_enrollment(enr, school_cct, cruce_por_cct, cruce_por_plan, raw_curp)
                    )
                    for warn in _advertencias_nombre(nombre, paterno, materno):
                        warnings.append(
                            {
                                "plantel": school_name,
                                "nombre": nombre,
                                "paterno": paterno,
                                "materno": materno,
                                "matricula": matricula,
                                "curp": raw_curp,
                                "warning": warn,
                            }
                        )
                except Exception as exc:
                    try:
                        resultados.append(
                            procesar_enrollment(
                                enr,
                                school_cct,
                                cruce_por_cct,
                                cruce_por_plan,
                                raw_curp,
                                skip_curp_validation=True,
                            )
                        )
                        warnings.append(
                            {
                                "plantel": school_name,
                                "nombre": nombre,
                                "paterno": paterno,
                                "materno": materno,
                                "matricula": matricula,
                                "curp": raw_curp,
                                "warning": str(exc),
                            }
                        )
                        for warn in _advertencias_nombre(nombre, paterno, materno):
                            warnings.append(
                                {
                                    "plantel": school_name,
                                    "nombre": nombre,
                                    "paterno": paterno,
                                    "materno": materno,
                                    "matricula": matricula,
                                    "curp": raw_curp,
                                    "warning": warn,
                                }
                            )
                    except Exception as exc2:
                        errores.append(
                            {
                                "plantel": school_name,
                                "nombre": nombre,
                                "paterno": paterno,
                                "materno": materno,
                                "matricula": matricula,
                                "curp": raw_curp,
                                "error": str(exc2),
                            }
                        )
            else:
                errores.append(
                    {
                        "plantel": school_name,
                        "nombre": nombre,
                        "paterno": paterno,
                        "materno": materno,
                        "matricula": matricula,
                        "curp": raw_curp,
                        "error": f"[CURPSuite] {suite_err}",
                    }
                )

        except Exception as exc:
            errores.append(
                {
                    "plantel": school_name,
                    "nombre": nombre,
                    "paterno": paterno,
                    "materno": materno,
                    "matricula": matricula,
                    "curp": raw_curp,
                    "error": str(exc),
                }
            )

        write_csv_rows(carga_path, resultados, DESTINO_FIELDNAMES)
        write_csv_rows(errores_path, errores, ERROR_FIELDNAMES)
        write_csv_rows(warnings_path, warnings, WARNING_FIELDNAMES)

        processed_matriculas.add(matricula)
        _guardar_checkpoint(
            timestamp,
            processed_matriculas,
            carga_path,
            errores_path,
            warnings_path,
            reporte_path,
        )

        total_ok += len(resultados)
        total_err += len(errores)
        total_warn += len(warnings)

    _eliminar_checkpoint()

    print(f"\nFinalizado {timestamp}")
    print(f"  Procesados : {total_ok}")
    print(f"  Warnings   : {total_warn}")
    print(f"  Errores    : {total_err}")
    if total_ok:
        print(f"  Carga CSV  : {carga_path}")
        carpeta_excel = csv_a_excel_particionado(carga_path)
        print(f"  Carga Excel: {carpeta_excel}")
    if total_warn:
        print(f"  Warnings   : {warnings_path}")
    if total_err:
        print(f"  Errores    : {errores_path}")


if __name__ == "__main__":
    main()
