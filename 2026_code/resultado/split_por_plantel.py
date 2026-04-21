import openpyxl
import os
import re
from collections import defaultdict

folder = r'D:\DEV\saekoAuth\2026_code\resultado\carga 2026-04-20 10-18-54\pendientes'
output_folder = os.path.join(folder, 'por_plantel')
os.makedirs(output_folder, exist_ok=True)

files = sorted([f for f in os.listdir(folder) if f.endswith('.xlsx') and not f.startswith('~$')])
print(f"Archivos encontrados: {files}")

plantel_rows = defaultdict(list)
headers = None

for filename in files:
    print(f"Leyendo {filename}...")
    wb = openpyxl.load_workbook(os.path.join(folder, filename), read_only=True, data_only=True)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            if headers is None:
                headers = list(row)
            continue
        plantel = str(row[2]).strip() if row[2] is not None else 'SIN_PLANTEL'
        plantel_rows[plantel].append(list(row))
    wb.close()
    print(f"  -> listo")

def safe_filename(name):
    return re.sub(r'[\\/*?:"<>|]', '_', name)

print(f"\nPlanteles encontrados: {len(plantel_rows)}")
for plantel, rows in sorted(plantel_rows.items()):
    fname = safe_filename(plantel) + '.xlsx'
    out_path = os.path.join(output_folder, fname)
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.append(headers)
    for row in rows:
        ws_out.append(row)
    wb_out.save(out_path)
    print(f"  Guardado: {fname} ({len(rows)} filas)")

print("\nListo.")
