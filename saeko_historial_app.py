"""
saeko_historial_app.py

Consulta un alumno por matrícula o CURP en la API de Saeko
y genera un Historial Académico en PDF incluyendo los créditos de cada materia.

Ejecutar: python saeko_historial_app.py
Requiere:  pip install reportlab
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import requests
import re
from datetime import datetime
from auth import SERVICE_ACCOUNT, USER_EMAIL, get_access_token

BASE_URL = "https://app.saeko.io/api/v1"

# ─────────────────────────────────────────────────────────────────────────────
# CAPA DE API
# ─────────────────────────────────────────────────────────────────────────────

def api_get(endpoint, access_token, params=None):
    headers = {"Authorization": f"Bearer {access_token}"}
    r = requests.get(
        f"{BASE_URL}{endpoint}", headers=headers, params=params, timeout=60
    )
    r.raise_for_status()
    return r.json()


def search_students(access_token, query):
    """
    Busca alumnos probando múltiples estrategias.
    Devuelve lista de dicts con al menos: id, full_name, curp.
    Lanza SearchError con detalles si ninguna estrategia da resultado.
    """
    query = query.strip()
    errors = []

    # ── Estrategia 1: /core/contacts?search= ─────────────────────────────────
    # Saeko almacena alumnos como contactos; este endpoint sí soporta search=.
    try:
        data = api_get("/core/contacts", access_token, {"search": query, "limit": 20})
        contacts = data.get("contacts", [])
        students = []
        for c in contacts:
            sub = c.get("student") or {}
            sid = sub.get("id") or c.get("student_id")
            if not sid:
                continue
            students.append({
                "id": sid,
                "full_name": (
                    c.get("full_name") or c.get("name") or
                    f"{c.get('last_name','')} {c.get('second_last_name','')} {c.get('first_name','')}".strip()
                ),
                "curp": c.get("curp") or sub.get("curp") or "",
                "current_enrollment": sub.get("current_enrollment") or {},
            })
        if students:
            return students
        if contacts:
            # Hubo contactos pero ninguno tiene student_id — informar
            errors.append(f"contacts: {len(contacts)} resultado(s) pero sin student_id")
    except requests.HTTPError as e:
        errors.append(f"contacts: HTTP {e.response.status_code}")
    except Exception as e:
        errors.append(f"contacts: {e}")

    # ── Estrategia 2: /core/students?search= (sin include_fields extra) ───────
    try:
        data = api_get("/core/students", access_token, {"search": query, "limit": 20})
        students = data.get("students") or []
        if students:
            return students
    except requests.HTTPError as e:
        errors.append(f"students: HTTP {e.response.status_code}")
    except Exception as e:
        errors.append(f"students: {e}")

    # ── Estrategia 3: enrollments con include student (para matrícula numérica)
    if re.match(r"^\d{5,}$", query):
        try:
            params = {
                "filters": f"enrollment_number={query}",
                "include_fields": "student",
                "limit": 5,
            }
            data = api_get("/core/enrollments", access_token, params)
            seen: set = set()
            students = []
            for e in data.get("enrollments", []):
                s = e.get("student") or {}
                sid = s.get("id")
                if sid and sid not in seen:
                    seen.add(sid)
                    students.append(s)
            if students:
                return students
        except requests.HTTPError as e:
            errors.append(f"enrollments: HTTP {e.response.status_code}")
        except Exception as e:
            errors.append(f"enrollments: {e}")

    raise SearchError(
        "No se encontraron alumnos.\n\n"
        f"Intentos realizados: {'; '.join(errors) or 'ninguno'}\n\n"
        "Prueba usar el campo 'ID Directo' con el ID numérico del alumno en Saeko."
    )


class SearchError(Exception):
    """Error de búsqueda sin resultado — no es un error de red."""
    pass


def get_student(access_token, student_id):
    params = {"include_fields": "curp,enrollment_number,addresses"}
    data = api_get(f"/core/students/{student_id}", access_token, params)
    return data.get("student", {})


def get_enrollments(access_token, student_id):
    """Obtiene todas las inscripciones del alumno (paginadas)."""
    all_items = []
    params = {
        "include_fields": (
            "term_name,group_name,school_name,program_name,"
            "grade_level,enrollment_number,control_number,is_current"
        ),
        "limit": 100,
        "offset": 0,
    }
    while True:
        data = api_get(f"/core/students/{student_id}/enrollments", access_token, params)
        items = data.get("enrollments", [])
        all_items.extend(items)
        if not data.get("meta", {}).get("next_page"):
            break
        params["offset"] += len(items)
    return all_items


def get_transcript(access_token, enrollment_id):
    """Obtiene el kardex/transcript de una inscripción."""
    params = {
        "include_fields": "transcript_records,earned_credits,total_credits,score_avg"
    }
    try:
        data = api_get(
            f"/certification/enrollments/{enrollment_id}/transcript",
            access_token,
            params,
        )
        return data.get("transcript", {})
    except Exception:
        return {}


def get_enrolled_courses(access_token, enrollment_id):
    """Calificaciones de un enrollment específico."""
    params = {
        "include_fields": (
            "score_ordinary,score_extraordinary,score_final,"
            "course_name,subject_id,subject_type"
        )
    }
    try:
        data = api_get(
            f"/grading/enrollments/{enrollment_id}/enrolled_courses",
            access_token,
            params,
        )
        return data.get("enrolled_courses", [])
    except Exception:
        return []


def get_all_student_enrolled_courses(access_token, student_id):
    """
    Obtiene TODAS las materias calificadas del alumno en todos los semestres
    usando el endpoint por estudiante (evita el problema de enrollments vacíos).
    Incluye enrollment_id para poder mapear a grade_level.
    """
    all_courses = []
    params = {
        "include_fields": (
            "course_name,subject_type,enrolled_type,"
            "score_ordinary,score_extraordinary,score_final,"
            "subject_id,enrollment_id"
        ),
        "limit": 500,
        "offset": 0,
    }
    try:
        while True:
            data = api_get(
                f"/grading/students/{student_id}/enrolled_courses",
                access_token,
                params,
            )
            courses = data.get("enrolled_courses", [])
            all_courses.extend(courses)
            if not data.get("meta", {}).get("next_page"):
                break
            params["offset"] += len(courses)
    except Exception:
        pass
    return all_courses


def get_program_subjects(access_token, program_id):
    """Materias del programa con sus créditos."""
    params = {"include_fields": "credits,grade_level,hours,mec_config"}
    try:
        data = api_get(f"/core/programs/{program_id}/subjects", access_token, params)
        return data.get("subjects", [])
    except Exception:
        return []


def get_school(access_token, school_id):
    try:
        data = api_get(f"/core/schools/{school_id}", access_token, {"include_fields": "cct,address"})
        return data.get("school", {})
    except Exception:
        return {}


# ─────────────────────────────────────────────────────────────────────────────
# PROCESAMIENTO DE DATOS
# ─────────────────────────────────────────────────────────────────────────────

def _extract_credits(subject: dict) -> int:
    """Extrae créditos de un registro de materia (intenta varios nombres de campo)."""
    mec = subject.get("mec_config")
    mec_credits = mec.get("credits") if isinstance(mec, dict) else None
    raw = (
        subject.get("credits")
        or subject.get("credit_hours")
        or mec_credits
        or subject.get("hours")
        or 0
    )
    try:
        return int(float(raw))
    except (TypeError, ValueError):
        return 0


def _score_type(ec: dict) -> str:
    """Infiere tipo de calificación (OR/EXT/ETS/CR/AC) desde enrolled_course."""
    if ec.get("score_extraordinary") is not None and ec.get("score_extraordinary") != "":
        return "EXT"
    return "OR"


def _compute_generation(enrollments: list) -> str:
    years = []
    for e in enrollments:
        for field in ("term_name", "begins_at"):
            m = re.search(r"\b(20\d{2})\b", str(e.get(field, "")))
            if m:
                years.append(int(m.group(1)))
                break
    if not years:
        return ""
    return f"{min(years)} - {max(years) + 1}"


def build_history(access_token, student_id, progress_cb=None):
    """
    Construye el historial académico completo del alumno.
    Devuelve un dict listo para pasarse a generate_pdf().
    """

    def progress(msg):
        if progress_cb:
            progress_cb(msg)

    progress("Obteniendo datos del alumno…")
    student = get_student(access_token, student_id)

    progress("Obteniendo inscripciones…")
    raw_enrollments = get_enrollments(access_token, student_id)

    # ── Agrupar TODAS las inscripciones por grade_level (sin descartar ninguna)
    # Un alumno puede tener múltiples enrollments por semestre (reinscripciones,
    # inscripciones de módulos separadas, etc.)
    by_grade_all: dict = {}   # grade_level → [enrollment, ...]
    enrollment_id_to_grade: dict = {}  # enrollment_id → grade_level
    for e in raw_enrollments:
        gl = e.get("grade_level") or 0
        enrollment_id_to_grade[e["id"]] = gl
        if gl not in by_grade_all:
            by_grade_all[gl] = []
        by_grade_all[gl].append(e)

    # Inscripción "canónica" por semestre: la más reciente (mayor ID) para metadata
    by_grade: dict = {
        gl: sorted(lst, key=lambda e: e.get("id", 0))[-1]
        for gl, lst in by_grade_all.items()
        if gl > 0
    }
    enrollments = sorted(by_grade.values(), key=lambda e: e.get("grade_level", 99))

    # ── Prefetch de TODOS los cursos calificados del alumno (1 sola llamada)
    progress("Cargando calificaciones globales del alumno…")
    all_student_courses = get_all_student_enrolled_courses(access_token, student_id)

    # Agrupar cursos globales por grade_level vía enrollment_id_to_grade
    global_by_grade: dict = {}   # grade_level → [enrolled_course, ...]
    for ec in all_student_courses:
        eid = ec.get("enrollment_id")
        gl = enrollment_id_to_grade.get(eid, 0) if eid else 0
        if gl > 0:
            global_by_grade.setdefault(gl, []).append(ec)

    # Datos de programa y plantel
    program_id = next(
        (e["program_id"] for e in enrollments if e.get("program_id")), None
    )
    school_id = next(
        (e["school_id"] for e in enrollments if e.get("school_id")), None
    )

    subject_credits: dict = {}
    if program_id:
        progress("Cargando créditos del programa…")
        for s in get_program_subjects(access_token, program_id):
            sid = s.get("id")
            if sid is not None:
                subject_credits[sid] = _extract_credits(s)

    school_data = {}
    if school_id:
        progress("Cargando datos del plantel…")
        school_data = get_school(access_token, school_id)

    # Construir semestres
    semesters = []
    all_scores = []   # score_float de todas las materias calificadas

    for enrollment in enrollments:
        eid = enrollment["id"]
        grade_level = enrollment.get("grade_level", 0)

        progress(f"Cargando semestre {grade_level}…")

        # ── 1. Intentar transcript de certificación (semestres finalizados) ──
        transcript = get_transcript(access_token, eid)
        records = transcript.get("transcript_records", [])

        if not records:
            # ── 2. Usar cursos del prefetch global (agrupa todos los enrollments
            #       del mismo semestre, por eso captura datos de inscripciones
            #       secundarias que el transcript individual no ve)
            ec_list = global_by_grade.get(grade_level, [])

            if not ec_list:
                # ── 3. Fallback: consultar TODOS los enrollments de este semestre
                for enroll_of_grade in by_grade_all.get(grade_level, []):
                    ec_list.extend(get_enrolled_courses(access_token, enroll_of_grade["id"]))

            # Deduplicar por nombre de materia (puede haber duplicados de varios enrollments)
            seen_names: set = set()
            for ec in ec_list:
                name = (ec.get("course_name") or "").strip()
                if not name or name in seen_names:
                    continue
                seen_names.add(name)
                score_raw = (
                    ec.get("score_final")
                    or ec.get("score_extraordinary")
                    or ec.get("score_ordinary")
                    or 0
                )
                try:
                    score = float(score_raw)
                except (TypeError, ValueError):
                    score = 0.0
                records.append(
                    {
                        "subject_name": name,
                        "subject_id": ec.get("subject_id"),
                        "score": score,
                        "score_type": _score_type(ec),
                        "credits": subject_credits.get(ec.get("subject_id"), 0),
                    }
                )
        else:
            # Enriquecer transcript_records con créditos del programa
            for rec in records:
                if "credits" not in rec or not rec["credits"]:
                    sid = rec.get("subject_id")
                    rec["credits"] = subject_credits.get(sid, 0)

        # Acumular calificaciones para el promedio general
        for rec in records:
            try:
                s = float(rec.get("score") or 0)
                if s > 0:
                    all_scores.append(s)
            except (TypeError, ValueError):
                pass

        semesters.append(
            {
                "grade_level": grade_level,
                "term_name": enrollment.get("term_name", ""),
                "group_name": enrollment.get("group_name", ""),
                "program_name": enrollment.get("program_name", ""),
                "school_name": enrollment.get("school_name", ""),
                "records": records,
            }
        )

    # ── Calcular créditos sumando directamente los registros ─────────────────
    # (el transcript de certificación frecuentemente está vacío)
    total_earned = 0
    total_possible = 0
    for sem in semesters:
        for rec in sem["records"]:
            cred = rec.get("credits") or 0
            try:
                cred = int(float(cred))
            except (TypeError, ValueError):
                cred = 0
            total_possible += cred
            try:
                score = float(rec.get("score") or 0)
            except (TypeError, ValueError):
                score = 0.0
            if score > 0:          # materia calificada = créditos cursados
                total_earned += cred

    overall_avg = round(sum(all_scores) / len(all_scores), 1) if all_scores else 0.0

    # ── Inscripción actual ────────────────────────────────────────────────────
    current_enroll = (
        next((e for e in enrollments if e.get("is_current")), None)
        or (enrollments[-1] if enrollments else {})
    )

    # ── Matrícula: buscar en los enrollments (enrollment_number / control_number)
    matricula_enroll = next(
        (
            e.get("enrollment_number") or e.get("control_number")
            for e in enrollments
            if e.get("enrollment_number") or e.get("control_number")
        ),
        "",
    )

    program_name = (
        current_enroll.get("program_name", "")
        or next((s["program_name"] for s in semesters if s.get("program_name")), "")
    )
    school_name = (
        school_data.get("name", "")
        or current_enroll.get("school_name", "")
        or next((s["school_name"] for s in semesters if s.get("school_name")), "")
    )

    return {
        "student": student,
        "school": school_data,
        "semesters": semesters,
        "program_name": program_name,
        "school_name": school_name,
        "school_cct": school_data.get("cct", ""),
        "current_term_name": current_enroll.get("term_name", ""),
        "current_grade_level": current_enroll.get("grade_level", ""),
        "current_group": current_enroll.get("group_name", ""),
        "generation": _compute_generation(enrollments),
        "total_earned_credits": total_earned,
        "total_credits": total_possible,
        "overall_avg": overall_avg,
        "matricula_enroll": matricula_enroll,   # matrícula del enrollment record
    }


# ─────────────────────────────────────────────────────────────────────────────
# GENERACIÓN DE PDF
# ─────────────────────────────────────────────────────────────────────────────

def _fmt_score(val) -> str:
    """Formatea una calificación: '6', '6.5', 'AC', etc."""
    if val is None or val == "" or val == 0:
        return "0"
    try:
        f = float(val)
        return str(int(f)) if f == int(f) else str(round(f, 1))
    except (TypeError, ValueError):
        return str(val)


def generate_pdf(history: dict, filepath: str):
    """Genera el historial académico en PDF usando reportlab."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm, mm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    except ImportError:
        raise RuntimeError(
            "Se requiere 'reportlab' para generar PDF.\n"
            "Instálalo con: pip install reportlab"
        )

    student = history["student"]
    school = history["school"]
    semesters = history["semesters"]

    # ── Nombres de campos (con fallbacks) ───────────────────────────────────
    student_name = (
        student.get("full_name")
        or f"{student.get('last_name', '')} {student.get('second_last_name', '')} {student.get('first_name', '')}".strip()
        or student.get("name", "")
    ).upper()
    curp = (student.get("curp") or "").upper()
    # Preferir matrícula del enrollment; si no, buscar en el student record
    matricula = (
        history.get("matricula_enroll")
        or student.get("enrollment_number")
        or student.get("student_number")
        or student.get("control_number")
        or str(student.get("id", ""))
    )
    school_name = history["school_name"]
    school_cct = history["school_cct"]
    program_name = history["program_name"]
    current_term = history["current_term_name"]
    current_grade = history["current_grade_level"]
    current_group = history["current_group"]
    generation = history["generation"]
    total_earned = history["total_earned_credits"]
    total_credits = history["total_credits"]
    overall_avg = history["overall_avg"]

    # Número a letras (para el promedio en el pie)
    def _num_to_words_simple(n):
        units = [
            "CERO","UNO","DOS","TRES","CUATRO","CINCO",
            "SEIS","SIETE","OCHO","NUEVE","DIEZ",
        ]
        if isinstance(n, (int, float)):
            ent = int(n)
            dec = round((n - ent) * 10)
            w = units[ent] if 0 <= ent <= 10 else str(ent)
            if dec:
                w += f" PUNTO {units[dec]}"
            return w
        return str(n)

    avg_words = _num_to_words_simple(overall_avg)
    _MESES = {
        1:"ENERO", 2:"FEBRERO", 3:"MARZO", 4:"ABRIL", 5:"MAYO", 6:"JUNIO",
        7:"JULIO", 8:"AGOSTO", 9:"SEPTIEMBRE", 10:"OCTUBRE",
        11:"NOVIEMBRE", 12:"DICIEMBRE",
    }
    _hoy = datetime.now()
    today_str = f"{_hoy.day} DE {_MESES[_hoy.month]} DEL {_hoy.year}"
    address = school.get("address", "") or ""

    # ── Estilos ───────────────────────────────────────────────────────────────
    W, H = A4  # 595.27 x 841.89 pts
    MARGIN = 1.5 * cm
    PAGE_W = W - 2 * MARGIN  # ~520 pts

    styles = getSampleStyleSheet()

    def ps(name, **kwargs):
        return ParagraphStyle(name, **kwargs)

    S_INST = ps("inst", fontSize=7, alignment=TA_CENTER, leading=9,
                fontName="Helvetica")
    S_INST_B = ps("instB", fontSize=8, alignment=TA_CENTER, leading=10,
                  fontName="Helvetica-Bold")
    S_TITLE = ps("title", fontSize=11, alignment=TA_CENTER, leading=14,
                 fontName="Helvetica-Bold")
    S_NORMAL = ps("norm", fontSize=7.5, alignment=TA_LEFT, leading=9.5,
                  fontName="Helvetica")
    S_BOLD = ps("bold", fontSize=7.5, alignment=TA_LEFT, leading=9.5,
                fontName="Helvetica-Bold")
    S_SMALL = ps("small", fontSize=6.5, alignment=TA_LEFT, leading=8,
                 fontName="Helvetica")
    S_FOOTER = ps("footer", fontSize=7, alignment=TA_CENTER, leading=9,
                  fontName="Helvetica")

    # ── Colores ───────────────────────────────────────────────────────────────
    C_HEADER = colors.HexColor("#1a3a5c")   # azul oscuro para cabecera
    C_SEM_H = colors.HexColor("#dce6f1")    # azul muy claro para encabezado semestre
    C_WHITE = colors.white
    C_LIGHT = colors.HexColor("#f5f5f5")
    C_BORDER = colors.HexColor("#aaaaaa")

    # ── Helpers de tabla ─────────────────────────────────────────────────────

    def sem_header(grade_level):
        ordinals = {
            1:"1º",2:"2º",3:"3º",4:"4º",5:"5º",6:"6º",7:"7º",8:"8º",
        }
        ord_str = ordinals.get(grade_level, f"{grade_level}°")
        return f"{ord_str} Semestre"

    def build_sem_table(sem_data, col_width):
        """Construye la tabla de calificaciones de un semestre."""
        if sem_data is None:
            # Semestre sin datos: tabla vacía con solo header
            gl = "—"
            rows = [[
                Paragraph(f"— Semestre", ps("sh", fontSize=7.5, fontName="Helvetica-Bold",
                                             alignment=TA_CENTER)),
                Paragraph("CALIF", ps("sh2", fontSize=7, fontName="Helvetica-Bold",
                                       alignment=TA_CENTER)),
                Paragraph("TIPO", ps("sh3", fontSize=7, fontName="Helvetica-Bold",
                                      alignment=TA_CENTER)),
                Paragraph("CRÉD", ps("sh4", fontSize=7, fontName="Helvetica-Bold",
                                      alignment=TA_CENTER)),
            ]]
            col_widths = [col_width * 0.6, col_width * 0.15, col_width * 0.12, col_width * 0.13]
            t = Table(rows, colWidths=col_widths)
            t.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), C_SEM_H),
                ("GRID", (0,0), (-1,-1), 0.4, C_BORDER),
                ("ALIGN", (1,0), (-1,-1), "CENTER"),
                ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
                ("BOTTOMPADDING", (0,0), (-1,-1), 3),
                ("TOPPADDING", (0,0), (-1,-1), 3),
            ]))
            return t

        grade_level = sem_data["grade_level"]
        records = sem_data["records"]

        sh = ps("sh", fontSize=7.5, fontName="Helvetica-Bold", alignment=TA_CENTER)
        sh2 = ps("sh2", fontSize=7, fontName="Helvetica-Bold", alignment=TA_CENTER)

        rows = [[
            Paragraph(sem_header(grade_level), sh),
            Paragraph("CALIF", sh2),
            Paragraph("TIPO", sh2),
            Paragraph("CRÉD", sh2),
        ]]

        sn_style = ps("sn", fontSize=6.5, fontName="Helvetica", alignment=TA_LEFT, leading=8)
        sc_style = ps("sc", fontSize=6.5, fontName="Helvetica", alignment=TA_CENTER, leading=8)

        for i, rec in enumerate(records):
            name = rec.get("subject_name") or rec.get("name") or ""
            score_raw = rec.get("score") or rec.get("final_score") or rec.get("grade") or 0
            score_type = (
                rec.get("score_type") or rec.get("status") or rec.get("grade_type") or ""
            ).upper()
            cred = rec.get("credits") or 0
            cred_str = str(int(cred)) if cred else "—"

            rows.append([
                Paragraph(name, sn_style),
                Paragraph(_fmt_score(score_raw), sc_style),
                Paragraph(score_type, sc_style),
                Paragraph(cred_str, sc_style),
            ])

        # Widths: name|grade|type|credits
        cw = [col_width * 0.60, col_width * 0.14, col_width * 0.12, col_width * 0.14]

        t = Table(rows, colWidths=cw, repeatRows=1)
        style = [
            # Header
            ("BACKGROUND", (0, 0), (-1, 0), C_SEM_H),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            # Grid
            ("GRID", (0, 0), (-1, -1), 0.4, C_BORDER),
            # Alternating rows
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_WHITE, C_LIGHT]),
            # Alignment
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 1), (0, -1), 3),
            ("RIGHTPADDING", (0, 1), (0, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]
        t.setStyle(TableStyle(style))
        return t

    # ── Construcción del documento ────────────────────────────────────────────
    doc = SimpleDocTemplate(
        filepath,
        pagesize=A4,
        leftMargin=MARGIN,
        rightMargin=MARGIN,
        topMargin=1.0 * cm,
        bottomMargin=1.5 * cm,
        title="Historial Académico",
    )

    story = []

    # ── 1. Encabezado institucional ──────────────────────────────────────────
    inst_lines = [
        Paragraph("COORDINACIÓN DE ORGANISMOS DESCENTRALIZADOS ESTATALES DE CECYTEs", S_INST),
        Paragraph("COLEGIO DE ESTUDIOS CIENTÍFICOS Y TECNOLÓGICOS DEL ESTADO DE MICHOACÁN DE OCAMPO", S_INST_B),
    ]
    if school_cct or school_name:
        inst_lines.append(
            Paragraph(
                f"CCT {school_cct}   {school_name}".strip().strip("-").strip(),
                S_INST,
            )
        )
    for p in inst_lines:
        story.append(p)

    story.append(Spacer(1, 4))
    story.append(Paragraph("HISTORIAL ACADÉMICO DEL ALUMNO", S_TITLE))
    story.append(HRFlowable(width="100%", thickness=1, color=C_HEADER, spaceAfter=4))

    # ── 2. Datos del alumno ───────────────────────────────────────────────────
    lbl = ps("lbl", fontSize=7.5, fontName="Helvetica-Bold", alignment=TA_LEFT, leading=10)
    val = ps("val", fontSize=7.5, fontName="Helvetica", alignment=TA_LEFT, leading=10)

    def row2(l1, v1, l2, v2):
        return [
            Paragraph(l1, lbl), Paragraph(v1, val),
            Paragraph(l2, lbl), Paragraph(v2, val),
        ]

    info_data = [
        row2("Alumno:", student_name, "Matrícula:", matricula),
        row2("CURP:", curp, "Periodo Actual:", current_term),
        row2("Semestre:", str(current_grade), "Grupo:", current_group),
        row2("Generación:", generation, "Bachillerato Tecnológico:", program_name),
    ]

    info_col_w = [PAGE_W * 0.12, PAGE_W * 0.38, PAGE_W * 0.16, PAGE_W * 0.34]
    info_table = Table(info_data, colWidths=info_col_w)
    info_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f0f4f8")),
        ("BOX", (0, 0), (-1, -1), 0.8, C_HEADER),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, C_BORDER),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 6))

    # ── 3. Tablas de calificaciones por semestre (pares lado a lado) ─────────
    # Organizar semestres en dict para acceso rápido
    sem_by_grade = {s["grade_level"]: s for s in semesters}

    # Calcular todos los semestres posibles (hasta el actual)
    max_grade = max((s["grade_level"] for s in semesters), default=6)
    max_grade = max(max_grade, 6)  # al menos 6 semestres

    # Presentar en pares (1-2, 3-4, 5-6)
    SEM_COL_W = (PAGE_W - 6) / 2  # 6 pts de separación

    for gl_left in range(1, max_grade + 1, 2):
        gl_right = gl_left + 1
        t_left = build_sem_table(sem_by_grade.get(gl_left), SEM_COL_W)
        t_right = build_sem_table(sem_by_grade.get(gl_right), SEM_COL_W) if gl_right <= max_grade else None

        if t_right:
            outer = Table(
                [[t_left, t_right]],
                colWidths=[SEM_COL_W + 3, SEM_COL_W + 3],
            )
        else:
            outer = Table(
                [[t_left, ""]],
                colWidths=[SEM_COL_W + 3, SEM_COL_W + 3],
            )

        outer.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(outer)
        story.append(Spacer(1, 5))

    # ── 4. Pie de página con créditos y promedio ──────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_BORDER, spaceBefore=4, spaceAfter=4))

    footer_text = (
        f"SE EXPIDE EL PRESENTE HISTORIAL ACADÉMICO CON <b>{total_earned} CRÉDITOS CURSADOS</b> "
        f"DE UN TOTAL DE {total_credits}, "
        f"CON UN PROMEDIO GENERAL DE <b>{overall_avg} ({avg_words})</b>, "
        f"EN {address}, "
        f"EL {today_str}, PARA LOS FINES DE CARÁCTER INFORMATIVO QUE AL INTERESADO CONVENGAN."
    )
    story.append(Paragraph(footer_text, S_FOOTER))
    story.append(Spacer(1, 14))

    # Líneas de firma
    sig_col_w = PAGE_W / 3
    sig_data = [[
        Paragraph("_" * 32, S_FOOTER),
        Paragraph("Sello<br/>plantel", S_FOOTER),
        Paragraph("_" * 32, S_FOOTER),
    ], [
        Paragraph("Director(a) del plantel", S_FOOTER),
        Paragraph("", S_FOOTER),
        Paragraph("Responsable de control escolar", S_FOOTER),
    ]]
    sig_table = Table(sig_data, colWidths=[sig_col_w, sig_col_w, sig_col_w])
    sig_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(sig_table)

    doc.build(story)


# ─────────────────────────────────────────────────────────────────────────────
# GENERACIÓN DE EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def generate_excel(history: dict, filepath: str):
    """Genera el historial académico en Excel (.xlsx) usando openpyxl."""
    import openpyxl
    from openpyxl.styles import (
        Font, Alignment, PatternFill, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    student    = history["student"]
    school     = history["school"]
    semesters  = history["semesters"]

    student_name = (
        student.get("full_name")
        or f"{student.get('last_name','')} {student.get('second_last_name','')} {student.get('first_name','')}".strip()
        or student.get("name", "")
    ).upper()
    curp         = (student.get("curp") or "").upper()
    matricula    = (
        history.get("matricula_enroll")
        or student.get("enrollment_number")
        or student.get("student_number")
        or student.get("control_number")
        or str(student.get("id", ""))
    )
    school_name   = history["school_name"]
    school_cct    = history["school_cct"]
    program_name  = history["program_name"]
    current_term  = history["current_term_name"]
    current_grade = history["current_grade_level"]
    current_group = history["current_group"]
    generation    = history["generation"]
    total_earned  = history["total_earned_credits"]
    total_credits = history["total_credits"]
    overall_avg   = history["overall_avg"]

    _MESES = {
        1:"ENERO",2:"FEBRERO",3:"MARZO",4:"ABRIL",5:"MAYO",6:"JUNIO",
        7:"JULIO",8:"AGOSTO",9:"SEPTIEMBRE",10:"OCTUBRE",
        11:"NOVIEMBRE",12:"DICIEMBRE",
    }
    _hoy = datetime.now()
    today_str = f"{_hoy.day} DE {_MESES[_hoy.month]} DEL {_hoy.year}"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial"

    # ── Estilos ───────────────────────────────────────────────────────────────
    def thin_border():
        s = Side(style="thin", color="AAAAAA")
        return Border(left=s, right=s, top=s, bottom=s)

    def thick_border():
        s = Side(style="medium", color="1A3A5C")
        return Border(left=s, right=s, top=s, bottom=s)

    C_DARK   = "1A3A5C"   # azul oscuro — encabezados de semestre
    C_LIGHT  = "DCE6F1"   # azul claro  — encabezado de semestre
    C_ALT    = "F5F5F5"   # gris claro  — fila alternada
    C_INFO   = "F0F4F8"   # azul muy claro — info alumno

    def style_header(cell, bg=C_DARK, bold=True, size=10, color="FFFFFF", wrap=False):
        cell.font      = Font(name="Calibri", bold=bold, size=size, color=color)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=wrap)
        cell.border    = thin_border()

    def style_label(cell):
        cell.font      = Font(name="Calibri", bold=True, size=9)
        cell.fill      = PatternFill("solid", fgColor=C_INFO)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border    = thin_border()

    def style_value(cell, bold=False):
        cell.font      = Font(name="Calibri", bold=bold, size=9)
        cell.fill      = PatternFill("solid", fgColor=C_INFO)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border    = thin_border()

    def style_subject(cell, alt=False):
        cell.font      = Font(name="Calibri", size=8)
        cell.fill      = PatternFill("solid", fgColor=C_ALT if alt else "FFFFFF")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border    = thin_border()

    def style_num(cell, alt=False, bold=False):
        cell.font      = Font(name="Calibri", size=8, bold=bold)
        cell.fill      = PatternFill("solid", fgColor=C_ALT if alt else "FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border()

    # ── Columnas: A=nombre materia, B=CALIF, C=TIPO, D=CRÉD ──────────────────
    #  Se usan 4 columnas por semestre, con un bloque a la vez (layout vertical)
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 7
    ws.column_dimensions["D"].width = 7

    row = 1

    # ── Encabezado institucional ──────────────────────────────────────────────
    inst_lines = [
        "COORDINACIÓN DE ORGANISMOS DESCENTRALIZADOS ESTATALES DE CECYTEs",
        "COLEGIO DE ESTUDIOS CIENTÍFICOS Y TECNOLÓGICOS DEL ESTADO DE MICHOACÁN DE OCAMPO",
    ]
    if school_cct or school_name:
        inst_lines.append(f"CCT {school_cct}   {school_name}".strip().strip("-").strip())

    for line in inst_lines:
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=line)
        c.font      = Font(name="Calibri", bold=(line == inst_lines[1]), size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 14
        row += 1

    # Título
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row=row, column=1, value="HISTORIAL ACADÉMICO DEL ALUMNO")
    c.font      = Font(name="Calibri", bold=True, size=13, color=C_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 20
    row += 1
    row += 1  # fila en blanco

    # ── Datos del alumno (tabla 2 columnas × 4 filas) ────────────────────────
    info_rows = [
        ("Alumno:",               student_name,  "Matrícula:",              matricula),
        ("CURP:",                 curp,          "Periodo Actual:",         current_term),
        ("Semestre:",             str(current_grade), "Grupo:",             current_group),
        ("Generación:",           generation,    "Bachillerato Tecnológico:", program_name),
    ]
    for lbl1, val1, lbl2, val2 in info_rows:
        # Col A = etiqueta, B-C merged = valor, D = etiqueta2, E-F sería val2
        # Con 4 columnas: A=lbl, B=val (merge BC), C=lbl2, D=val2
        ws.merge_cells(f"B{row}:B{row}")   # no merge needed actually
        c_l1 = ws.cell(row=row, column=1, value=lbl1)
        c_v1 = ws.cell(row=row, column=2, value=val1)
        # put label2 + val2 in C and D — but we only have 4 cols, so we'll expand
        # Actually use cols A,B,C,D with widths: label narrow, value wide
        # Let's just use 4 cols: A(lbl) B(val) C(lbl2) D(val2)
        c_l2 = ws.cell(row=row, column=3, value=lbl2)
        c_v2 = ws.cell(row=row, column=4, value=val2)
        style_label(c_l1); style_value(c_v1)
        style_label(c_l2); style_value(c_v2)
        ws.row_dimensions[row].height = 15
        row += 1

    row += 1  # separador

    # ── Tablas de calificaciones por semestre ─────────────────────────────────
    ordinals = {1:"1º",2:"2º",3:"3º",4:"4º",5:"5º",6:"6º",7:"7º",8:"8º"}

    for sem in semesters:
        gl      = sem["grade_level"]
        records = sem["records"]
        ord_str = ordinals.get(gl, f"{gl}°")

        # Encabezado semestre
        ws.merge_cells(f"A{row}:A{row}")
        c_sh = ws.cell(row=row, column=1, value=f"{ord_str} Semestre")
        style_header(c_sh, bg=C_LIGHT, color=C_DARK, size=9, bold=True)
        c_calif = ws.cell(row=row, column=2, value="CALIF")
        style_header(c_calif, bg=C_LIGHT, color=C_DARK, size=8)
        c_tipo  = ws.cell(row=row, column=3, value="TIPO")
        style_header(c_tipo,  bg=C_LIGHT, color=C_DARK, size=8)
        c_cred  = ws.cell(row=row, column=4, value="CRÉD")
        style_header(c_cred,  bg=C_LIGHT, color=C_DARK, size=8)
        ws.row_dimensions[row].height = 14
        row += 1

        if not records:
            ws.merge_cells(f"A{row}:D{row}")
            c = ws.cell(row=row, column=1, value="(Sin registros)")
            c.font      = Font(name="Calibri", italic=True, size=8, color="888888")
            c.alignment = Alignment(horizontal="center")
            c.border    = thin_border()
            ws.row_dimensions[row].height = 12
            row += 1
        else:
            for i, rec in enumerate(records):
                alt  = (i % 2 == 1)
                name = rec.get("subject_name") or rec.get("name") or ""
                score_raw = rec.get("score") or rec.get("final_score") or rec.get("grade") or 0
                stype = (
                    rec.get("score_type") or rec.get("status") or rec.get("grade_type") or ""
                ).upper()
                cred = rec.get("credits") or 0
                try:
                    cred = int(float(cred))
                except (TypeError, ValueError):
                    cred = 0

                # Formatear calificación
                try:
                    score_f = float(score_raw)
                    score_disp = int(score_f) if score_f == int(score_f) else round(score_f, 1)
                except (TypeError, ValueError):
                    score_disp = score_raw

                c_n = ws.cell(row=row, column=1, value=name)
                c_s = ws.cell(row=row, column=2, value=score_disp)
                c_t = ws.cell(row=row, column=3, value=stype)
                c_c = ws.cell(row=row, column=4, value=cred if cred else "—")

                style_subject(c_n, alt)
                style_num(c_s, alt)
                style_num(c_t, alt)
                style_num(c_c, alt)
                ws.row_dimensions[row].height = 13
                row += 1

        row += 1  # espacio entre semestres

    # ── Pie: créditos y promedio ───────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    footer = (
        f"SE EXPIDE EL PRESENTE HISTORIAL ACADÉMICO CON {total_earned} CRÉDITOS CURSADOS "
        f"DE UN TOTAL DE {total_credits}, CON UN PROMEDIO GENERAL DE {overall_avg}, "
        f"EL {today_str}."
    )
    c_f = ws.cell(row=row, column=1, value=footer)
    c_f.font      = Font(name="Calibri", size=8, italic=True)
    c_f.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c_f.border    = thin_border()
    ws.row_dimensions[row].height = 28

    # Ajustar altura de filas de info
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 28

    wb.save(filepath)


# ─────────────────────────────────────────────────────────────────────────────
# INTERFAZ GRÁFICA (TKINTER)
# ─────────────────────────────────────────────────────────────────────────────

class SaekoHistorialApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Saeko — Historial Académico con Créditos")
        self.root.geometry("820x640")
        self.root.resizable(True, True)

        self.access_token = None
        self._students = []       # resultados de búsqueda
        self._history = None      # historial cargado del alumno seleccionado

        self._build_ui()
        self._authenticate()

    # ── UI ────────────────────────────────────────────────────────────────────

    def _build_ui(self):
        main = ttk.Frame(self.root, padding=12)
        main.pack(fill=tk.BOTH, expand=True)

        ttk.Label(
            main,
            text="Historial Académico — Consulta por Matrícula o CURP",
            font=("Segoe UI", 13, "bold"),
        ).pack(pady=(0, 8))

        # ── Búsqueda ──
        search_frame = ttk.LabelFrame(main, text="Buscar alumno", padding=8)
        search_frame.pack(fill=tk.X, pady=(0, 8))

        row = ttk.Frame(search_frame)
        row.pack(fill=tk.X, pady=(0, 4))
        ttk.Label(row, text="Nombre / CURP / Matrícula:", width=26, anchor="w").pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(row, textvariable=self.search_var, width=36)
        self.search_entry.pack(side=tk.LEFT, padx=(0, 8))
        self.search_entry.bind("<Return>", lambda _: self._on_search())
        self.btn_search = ttk.Button(row, text="Buscar", command=self._on_search, width=10)
        self.btn_search.pack(side=tk.LEFT)

        row2 = ttk.Frame(search_frame)
        row2.pack(fill=tk.X)
        ttk.Label(row2, text="— o bien —  ID de Saeko:", width=26, anchor="w",
                  foreground="#666").pack(side=tk.LEFT)
        self.id_var = tk.StringVar()
        id_entry = ttk.Entry(row2, textvariable=self.id_var, width=14)
        id_entry.pack(side=tk.LEFT, padx=(0, 8))
        id_entry.bind("<Return>", lambda _: self._on_load_by_id())
        ttk.Button(row2, text="Cargar por ID", command=self._on_load_by_id, width=14).pack(side=tk.LEFT)
        ttk.Label(row2, text="(ID interno de Saeko, sin búsqueda)", foreground="#999",
                  font=("Segoe UI", 8)).pack(side=tk.LEFT, padx=(8, 0))

        # ── Lista de resultados ──
        res_frame = ttk.LabelFrame(main, text="Resultados", padding=6)
        res_frame.pack(fill=tk.X, pady=(0, 8))

        cols = ("id", "nombre", "curp", "programa", "semestre")
        self.tree_results = ttk.Treeview(
            res_frame, columns=cols, show="headings", height=5, selectmode="browse"
        )
        self.tree_results.heading("id", text="ID")
        self.tree_results.heading("nombre", text="Nombre")
        self.tree_results.heading("curp", text="CURP")
        self.tree_results.heading("programa", text="Programa")
        self.tree_results.heading("semestre", text="Semestre")
        self.tree_results.column("id", width=55, anchor="center")
        self.tree_results.column("nombre", width=230, anchor="w")
        self.tree_results.column("curp", width=145, anchor="center")
        self.tree_results.column("programa", width=190, anchor="w")
        self.tree_results.column("semestre", width=70, anchor="center")
        vsb = ttk.Scrollbar(res_frame, orient="vertical", command=self.tree_results.yview)
        self.tree_results.configure(yscrollcommand=vsb.set)
        self.tree_results.pack(side=tk.LEFT, fill=tk.X, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree_results.bind("<<TreeviewSelect>>", self._on_select_student)

        # ── Info del alumno seleccionado ──
        info_frame = ttk.LabelFrame(main, text="Datos del alumno seleccionado", padding=8)
        info_frame.pack(fill=tk.X, pady=(0, 8))

        self.info_text = tk.Text(info_frame, height=5, state="disabled",
                                 font=("Consolas", 8), wrap="word",
                                 background="#f8f8f8")
        self.info_text.pack(fill=tk.X)

        # ── Botones de acción ──
        btn_frame = ttk.Frame(main)
        btn_frame.pack(fill=tk.X, pady=(0, 6))

        self.btn_load = ttk.Button(
            btn_frame, text="Cargar historial completo",
            command=self._on_load_history, state="disabled", width=26
        )
        self.btn_load.pack(side=tk.LEFT, padx=(0, 10))

        self.btn_pdf = ttk.Button(
            btn_frame, text="Guardar PDF",
            command=self._on_generate_pdf, state="disabled", width=16
        )
        self.btn_pdf.pack(side=tk.LEFT, padx=(0, 6))

        self.btn_excel = ttk.Button(
            btn_frame, text="Guardar Excel",
            command=self._on_generate_excel, state="disabled", width=16
        )
        self.btn_excel.pack(side=tk.LEFT)

        # ── Progreso y estado ──
        self.progress = ttk.Progressbar(main, mode="indeterminate")
        self.progress.pack(fill=tk.X, pady=(0, 3))

        self.status_var = tk.StringVar(value="Iniciando…")
        ttk.Label(main, textvariable=self.status_var,
                  font=("Segoe UI", 8), foreground="#555").pack(anchor="w")

    # ── Autenticación ─────────────────────────────────────────────────────────

    def _authenticate(self):
        self.status_var.set("Autenticando con Saeko…")
        self.btn_search.configure(state="disabled")

        def task():
            try:
                token_data = get_access_token(SERVICE_ACCOUNT, USER_EMAIL)
                self.access_token = token_data["access_token"]
                self.root.after(0, lambda: self.status_var.set(
                    "Listo. Ingresa una matrícula o CURP y presiona Buscar."
                ))
                self.root.after(0, lambda: self.btn_search.configure(state="normal"))
                self.root.after(0, lambda: self.search_entry.focus_set())
            except Exception as e:
                msg = str(e)
                self.root.after(
                    0, lambda: messagebox.showerror("Error de autenticación", msg)
                )

        threading.Thread(target=task, daemon=True).start()

    # ── Búsqueda ──────────────────────────────────────────────────────────────

    def _on_search(self):
        query = self.search_var.get().strip()
        if not query:
            messagebox.showwarning("Atención", "Ingresa nombre, CURP o matrícula.")
            return

        self._set_busy(True)
        self.status_var.set(f"Buscando '{query}'…")
        for row in self.tree_results.get_children():
            self.tree_results.delete(row)
        self._history = None
        self.btn_pdf.configure(state="disabled")
        self.btn_excel.configure(state="disabled")
        self._update_info("")

        def task():
            try:
                students = search_students(self.access_token, query)
                self.root.after(0, lambda s=students: self._populate_results(s))
            except SearchError as e:
                msg = str(e)
                self.root.after(0, lambda m=msg: (
                    self.status_var.set("Sin resultados. Usa 'ID de Saeko' si conoces el ID."),
                    messagebox.showinfo("Sin resultados", m),
                ))
            except Exception as e:
                msg = str(e)
                self.root.after(0, lambda m=msg: (
                    self.status_var.set(f"Error: {m[:80]}"),
                    messagebox.showerror("Error de búsqueda", m),
                ))
            finally:
                self.root.after(0, lambda: self._set_busy(False))

        threading.Thread(target=task, daemon=True).start()

    def _on_load_by_id(self):
        """Carga directamente un alumno por su ID interno de Saeko."""
        raw = self.id_var.get().strip()
        if not raw or not raw.isdigit():
            messagebox.showwarning("Atención", "Ingresa un ID numérico de Saeko.")
            return
        student_id = int(raw)

        self._set_busy(True)
        self.status_var.set(f"Cargando alumno ID {student_id}…")
        for row in self.tree_results.get_children():
            self.tree_results.delete(row)
        self._history = None
        self.btn_pdf.configure(state="disabled")
        self.btn_excel.configure(state="disabled")
        self._update_info("")

        def task():
            try:
                student = get_student(self.access_token, student_id)
                if not student:
                    raise ValueError(f"No se encontró alumno con ID {student_id}.")
                self.root.after(0, lambda s=student: self._populate_results([s]))
            except Exception as e:
                msg = str(e)
                self.root.after(0, lambda m=msg: (
                    self.status_var.set(f"Error: {m[:80]}"),
                    messagebox.showerror("Error", m),
                ))
            finally:
                self.root.after(0, lambda: self._set_busy(False))

        threading.Thread(target=task, daemon=True).start()

    def _populate_results(self, students):
        self._students = students
        for row in self.tree_results.get_children():
            self.tree_results.delete(row)

        if not students:
            self.status_var.set("No se encontraron alumnos.")
            return

        for s in students:
            name = (
                s.get("full_name")
                or f"{s.get('last_name','')} {s.get('second_last_name','')} {s.get('first_name','')}".strip()
                or s.get("name", "")
            )
            curp = (s.get("curp") or "").upper()
            enroll = s.get("current_enrollment") or {}
            prog = enroll.get("program_name", "")
            sem = enroll.get("grade_level", "")

            self.tree_results.insert(
                "", tk.END,
                iid=str(s["id"]),
                values=(s["id"], name.upper(), curp, prog, sem),
            )

        self.status_var.set(
            f"{len(students)} alumno(s) encontrado(s). Selecciona uno para cargar su historial."
        )

    # ── Selección de alumno ───────────────────────────────────────────────────

    def _on_select_student(self, event=None):
        sel = self.tree_results.selection()
        if not sel:
            return
        student_id = int(sel[0])
        student = next((s for s in self._students if s["id"] == student_id), None)
        if not student:
            return

        name = (
            student.get("full_name")
            or f"{student.get('last_name','')} {student.get('second_last_name','')} {student.get('first_name','')}".strip()
            or student.get("name", "")
        )
        curp = (student.get("curp") or "").upper()
        enroll = student.get("current_enrollment") or {}

        info = (
            f"ID:         {student_id}\n"
            f"Nombre:     {name.upper()}\n"
            f"CURP:       {curp}\n"
            f"Programa:   {enroll.get('program_name','')}\n"
            f"Semestre:   {enroll.get('grade_level','')}"
        )
        self._update_info(info)
        self.btn_load.configure(state="normal")
        self._history = None
        self.btn_pdf.configure(state="disabled")
        self.btn_excel.configure(state="disabled")

    # ── Cargar historial ──────────────────────────────────────────────────────

    def _on_load_history(self):
        sel = self.tree_results.selection()
        if not sel:
            return
        student_id = int(sel[0])

        self._set_busy(True)
        self.btn_pdf.configure(state="disabled")
        self.btn_excel.configure(state="disabled")
        self._history = None

        def task():
            try:
                history = build_history(
                    self.access_token,
                    student_id,
                    progress_cb=lambda msg: self.root.after(
                        0, lambda m=msg: self.status_var.set(m)
                    ),
                )
                self._history = history
                self.root.after(0, lambda: self._show_history_summary())
            except Exception as e:
                msg = str(e)
                self.root.after(
                    0,
                    lambda: (
                        self.status_var.set(f"Error: {msg[:80]}"),
                        messagebox.showerror("Error al cargar historial", msg),
                    ),
                )
            finally:
                self.root.after(0, lambda: self._set_busy(False))

        threading.Thread(target=task, daemon=True).start()

    def _show_history_summary(self):
        h = self._history
        sems = h.get("semesters", [])
        total_mats = sum(len(s["records"]) for s in sems)
        lines = [
            f"Alumno:       {h['student'].get('full_name','') or h['student'].get('name','')}",
            f"Programa:     {h['program_name']}",
            f"Plantel:      {h['school_name']}",
            f"Semestres:    {len(sems)} cargados | {total_mats} materias en total",
            f"Créditos:     {h['total_earned_credits']} / {h['total_credits']}",
            f"Promedio:     {h['overall_avg']}",
        ]
        self._update_info("\n".join(lines))
        self.status_var.set(
            f"Historial cargado: {len(sems)} semestres, {total_mats} materias. "
            "Usa 'Guardar PDF' o 'Guardar Excel'."
        )
        self.btn_pdf.configure(state="normal")
        self.btn_excel.configure(state="normal")

    # ── Generar PDF ───────────────────────────────────────────────────────────

    def _on_generate_pdf(self):
        if not self._history:
            messagebox.showwarning("Atención", "Primero carga el historial del alumno.")
            return

        student = self._history["student"]
        student_name = (
            student.get("full_name")
            or f"{student.get('last_name','')} {student.get('first_name','')}".strip()
            or "alumno"
        )
        default_name = f"historial_{student_name.replace(' ', '_').lower()}.pdf"

        filepath = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf"), ("All files", "*.*")],
            title="Guardar historial académico como…",
            initialfile=default_name,
        )
        if not filepath:
            return

        self._set_busy(True)
        self.status_var.set("Generando PDF…")

        def task():
            try:
                generate_pdf(self._history, filepath)
                self.root.after(
                    0,
                    lambda: (
                        self.status_var.set(f"PDF guardado: {filepath}"),
                        messagebox.showinfo(
                            "Listo",
                            f"Historial académico generado exitosamente.\n\n{filepath}",
                        ),
                    ),
                )
            except RuntimeError as e:
                msg = str(e)
                self.root.after(
                    0,
                    lambda: messagebox.showerror("Módulo faltante", msg),
                )
            except Exception as e:
                msg = str(e)
                self.root.after(
                    0,
                    lambda: messagebox.showerror("Error al generar PDF", msg),
                )
            finally:
                self.root.after(0, lambda: self._set_busy(False))

        threading.Thread(target=task, daemon=True).start()

    # ── Generar Excel ─────────────────────────────────────────────────────────

    def _on_generate_excel(self):
        if not self._history:
            messagebox.showwarning("Atención", "Primero carga el historial del alumno.")
            return

        student = self._history["student"]
        student_name = (
            student.get("full_name")
            or f"{student.get('last_name','')} {student.get('first_name','')}".strip()
            or "alumno"
        )
        default_name = f"historial_{student_name.replace(' ', '_').lower()}.xlsx"

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("All files", "*.*")],
            title="Guardar historial como Excel…",
            initialfile=default_name,
        )
        if not filepath:
            return

        self._set_busy(True)
        self.status_var.set("Generando Excel…")

        def task():
            try:
                generate_excel(self._history, filepath)
                self.root.after(0, lambda: (
                    self.status_var.set(f"Excel guardado: {filepath}"),
                    messagebox.showinfo("Listo",
                        f"Historial exportado a Excel.\n\n{filepath}"),
                ))
            except Exception as e:
                msg = str(e)
                self.root.after(0, lambda m=msg: messagebox.showerror(
                    "Error al generar Excel", m))
            finally:
                self.root.after(0, lambda: self._set_busy(False))

        threading.Thread(target=task, daemon=True).start()

    # ── Helpers UI ────────────────────────────────────────────────────────────

    def _set_busy(self, busy: bool):
        if busy:
            self.progress.start(10)
            self.btn_search.configure(state="disabled")
            self.btn_load.configure(state="disabled")
        else:
            self.progress.stop()
            self.btn_search.configure(state="normal")
            if self.tree_results.selection():
                self.btn_load.configure(state="normal")
            if self._history:
                self.btn_pdf.configure(state="normal")
                self.btn_excel.configure(state="normal")

    def _update_info(self, text: str):
        self.info_text.configure(state="normal")
        self.info_text.delete("1.0", tk.END)
        self.info_text.insert("1.0", text)
        self.info_text.configure(state="disabled")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    root = tk.Tk()
    app = SaekoHistorialApp(root)
    root.mainloop()
