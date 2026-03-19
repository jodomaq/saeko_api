"""
Saeko Nuevo Ingreso App
Aplicación Tkinter para consultar el total de alumnos de PRIMER SEMESTRE (nuevo ingreso)
por plantel, filtrando únicamente los periodos que inician en agosto.
Genera una matriz Plantel × Periodo con exportación a Excel.
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import requests
from auth import SERVICE_ACCOUNT, USER_EMAIL, get_access_token

BASE_URL = "https://app.saeko.io/api/v1"


# ============================================================
# CAPA DE API
# ============================================================

def api_get(endpoint: str, access_token: str, params: dict = None) -> dict:
    headers = {"Authorization": f"Bearer {access_token}"}
    response = requests.get(f"{BASE_URL}{endpoint}", headers=headers, params=params, timeout=60)
    response.raise_for_status()
    return response.json()


def get_schools(access_token: str) -> list:
    data = api_get("/core/schools", access_token, {"include_fields": "cct"})
    return data.get("schools", [])


def get_terms_by_school(access_token: str, school_id: int) -> list:
    """Obtiene todos los términos de un plantel manejando paginación."""
    all_terms = []
    params = {"limit": 200, "offset": 0}
    while True:
        data = api_get(f"/core/schools/{school_id}/terms", access_token, params)
        terms = data.get("terms", [])
        all_terms.extend(terms)
        if not data.get("meta", {}).get("next_page"):
            break
        params["offset"] += len(terms)
    return all_terms


def is_august_term(term_name: str) -> bool:
    """Devuelve True si el nombre del periodo contiene 'agosto' (mayúsculas o no).

    Se usa el nombre en lugar de begins_at porque los periodos llamados 'Agosto'
    pueden tener begins_at en septiembre (fecha real de inicio de clases).
    """
    return "agosto" in term_name.lower()


def get_nuevo_ingreso_count(access_token: str, school_id: int, term_id: int) -> int:
    """
    Obtiene el total de alumnos de primer semestre (grade_level=1) de un plantel
    en un term usando limit=1 y leyendo meta.total.
    """
    params = {
        "limit": 1,
        "offset": 0,
        "filters": f"school_id={school_id};grade_level=1",
    }
    data = api_get(f"/core/terms/{term_id}/enrollments", access_token, params)
    return data.get("meta", {}).get("total", len(data.get("enrollments", [])))


# ============================================================
# APLICACIÓN TKINTER
# ============================================================

class SaekoNuevoIngresoApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Saeko - Nuevo Ingreso (1er Semestre) × Plantel × Agosto")
        self.root.geometry("1100x680")
        self.root.resizable(True, True)

        self.access_token = None
        self.schools = []
        self._cancel_event = threading.Event()
        self._sorted_term_names = []   # lista ordenada de nombres de periodos de agosto
        self._term_info = {}           # term_name → {begins_at, ends_at, is_current}
        self._school_term_map = {}     # school_id → {term_name → term_id}
        self.selected_term_names = []
        self.matrix = {}               # (school_id, term_name) → count

        self._build_ui()
        self._authenticate()

    def _build_ui(self):
        main = ttk.Frame(self.root, padding=15)
        main.pack(fill=tk.BOTH, expand=True)

        ttk.Label(
            main,
            text="Nuevo Ingreso — Alumnos de 1er Semestre (Periodos Agosto)",
            font=("Segoe UI", 15, "bold"),
        ).pack(pady=(0, 4))

        ttk.Label(
            main,
            text="Solo se muestran periodos que inician en agosto. Solo cuenta inscripciones de grade_level=1.",
            font=("Segoe UI", 9),
            foreground="#555555",
        ).pack(pady=(0, 12))

        # --- Filtros de rango de periodos ---
        filter_frame = ttk.LabelFrame(main, text="Rango de Periodos (solo agosto)", padding=10)
        filter_frame.pack(fill=tk.X, pady=(0, 10))

        row0 = ttk.Frame(filter_frame)
        row0.pack(fill=tk.X, pady=3)
        ttk.Label(row0, text="Periodo Inicio:", width=18, anchor="w").pack(side=tk.LEFT)
        self.term_start_var = tk.StringVar()
        self.term_start_combo = ttk.Combobox(row0, textvariable=self.term_start_var,
                                              state="readonly", width=65)
        self.term_start_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)

        row1 = ttk.Frame(filter_frame)
        row1.pack(fill=tk.X, pady=3)
        ttk.Label(row1, text="Periodo Fin:", width=18, anchor="w").pack(side=tk.LEFT)
        self.term_end_var = tk.StringVar()
        self.term_end_combo = ttk.Combobox(row1, textvariable=self.term_end_var,
                                            state="readonly", width=65)
        self.term_end_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # --- Botones ---
        btn_frame = ttk.Frame(main)
        btn_frame.pack(fill=tk.X, pady=(5, 10))
        self.btn_query = ttk.Button(btn_frame, text="Consultar Nuevo Ingreso",
                                     command=self._on_query)
        self.btn_query.pack(side=tk.LEFT, padx=(0, 10))
        self.btn_export = ttk.Button(btn_frame, text="Exportar a Excel",
                                      command=self._on_export, state="disabled")
        self.btn_export.pack(side=tk.LEFT, padx=(0, 10))
        self.btn_cancel = ttk.Button(btn_frame, text="Cancelar",
                                      command=self._on_cancel, state="disabled")
        self.btn_cancel.pack(side=tk.LEFT)

        # --- Barra de progreso ---
        self.progress = ttk.Progressbar(main, mode="determinate")
        self.progress.pack(fill=tk.X, pady=(0, 5))

        self.status_var = tk.StringVar(value="Iniciando...")
        ttk.Label(main, textvariable=self.status_var,
                  font=("Segoe UI", 9)).pack(anchor="w")

        # --- Tabla de resultados ---
        self.tree_frame = ttk.Frame(main)
        self.tree_frame.pack(fill=tk.BOTH, expand=True, pady=(5, 0))

        self.tree = None
        self._build_empty_tree()

    def _build_empty_tree(self):
        if self.tree:
            self.tree.destroy()
        for w in self.tree_frame.winfo_children():
            w.destroy()

        columns = ("no", "cct", "nombre_plantel")
        self.tree = ttk.Treeview(self.tree_frame, columns=columns, show="headings", height=18)
        self.tree.heading("no", text="#")
        self.tree.heading("cct", text="CCT")
        self.tree.heading("nombre_plantel", text="Nombre del Plantel")
        self.tree.column("no", width=40, anchor="center")
        self.tree.column("cct", width=130, anchor="center")
        self.tree.column("nombre_plantel", width=400, anchor="w")

        vsb = ttk.Scrollbar(self.tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(self.tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.tree_frame.grid_rowconfigure(0, weight=1)
        self.tree_frame.grid_columnconfigure(0, weight=1)

    def _build_matrix_tree(self, term_names):
        for w in self.tree_frame.winfo_children():
            w.destroy()

        col_ids = ["no", "cct", "nombre_plantel"]
        for idx in range(len(term_names)):
            col_ids.append(f"term_{idx}")
        col_ids.append("total")

        self.tree = ttk.Treeview(self.tree_frame, columns=col_ids, show="headings", height=18)

        self.tree.heading("no", text="#")
        self.tree.column("no", width=35, anchor="center", minwidth=35)
        self.tree.heading("cct", text="CCT")
        self.tree.column("cct", width=120, anchor="center", minwidth=80)
        self.tree.heading("nombre_plantel", text="Nombre del Plantel")
        self.tree.column("nombre_plantel", width=280, anchor="w", minwidth=150)

        for idx, name in enumerate(term_names):
            col_id = f"term_{idx}"
            self.tree.heading(col_id, text=name)
            self.tree.column(col_id, width=90, anchor="center", minwidth=60)

        self.tree.heading("total", text="TOTAL")
        self.tree.column("total", width=80, anchor="center", minwidth=60)

        vsb = ttk.Scrollbar(self.tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(self.tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.tree_frame.grid_rowconfigure(0, weight=1)
        self.tree_frame.grid_columnconfigure(0, weight=1)

    # --- Autenticación ---
    def _authenticate(self):
        self.status_var.set("Autenticando...")
        self._disable_controls()

        def task():
            try:
                token_data = get_access_token(SERVICE_ACCOUNT, USER_EMAIL)
                self.access_token = token_data["access_token"]
                self.root.after(0, self._load_schools_and_terms)
            except Exception as e:
                msg = str(e)
                self.root.after(0, lambda msg=msg: self._show_error("Error de autenticación", msg))

        threading.Thread(target=task, daemon=True).start()

    def _load_schools_and_terms(self):
        self.status_var.set("Cargando planteles y periodos de agosto...")

        def task():
            try:
                self.schools = get_schools(self.access_token)
                self._school_term_map = {}
                self._term_info = {}

                for i, school in enumerate(self.schools):
                    school_id = school["id"]
                    self.root.after(0, lambda i=i: self.status_var.set(
                        f"Cargando periodos: plantel {i + 1}/{len(self.schools)}..."
                    ))
                    try:
                        terms = get_terms_by_school(self.access_token, school_id)
                        school_map = {}
                        for t in terms:
                            # Solo periodos cuyo nombre contiene "agosto"
                            term_name = t.get("name", f"Term {t['id']}")
                            if not is_august_term(term_name):
                                continue
                            school_map[term_name] = t["id"]
                            if term_name not in self._term_info:
                                self._term_info[term_name] = {
                                    "begins_at": t.get("begins_at", ""),
                                    "ends_at": t.get("ends_at", ""),
                                    "is_current": t.get("is_current", False),
                                }
                            elif t.get("is_current"):
                                self._term_info[term_name]["is_current"] = True
                        self._school_term_map[school_id] = school_map
                    except Exception:
                        self._school_term_map[school_id] = {}

                self.root.after(0, self._populate_terms)
            except Exception as e:
                msg = str(e)
                self.root.after(0, lambda msg=msg: self._show_error("Error cargando datos", msg))

        threading.Thread(target=task, daemon=True).start()

    def _populate_terms(self):
        sorted_names = sorted(
            self._term_info.keys(),
            key=lambda n: self._term_info[n].get("begins_at", "")
        )
        self._sorted_term_names = sorted_names

        if not sorted_names:
            self.status_var.set(
                "No se encontraron periodos que inicien en agosto. "
                "Verifica los datos del sistema."
            )
            self._enable_controls()
            return

        term_values = []
        for name in sorted_names:
            info = self._term_info[name]
            current = " [ACTUAL]" if info.get("is_current") else ""
            term_values.append(
                f"{name} | {info.get('begins_at', '')} → {info.get('ends_at', '')}{current}"
            )
        self.term_start_combo["values"] = term_values
        self.term_end_combo["values"] = term_values

        # Seleccionar el periodo actual (o el último) como fin
        current_idx = None
        for i, name in enumerate(sorted_names):
            if self._term_info[name].get("is_current"):
                current_idx = i
                break

        if current_idx is not None:
            self.term_end_combo.current(current_idx)
            start_idx = max(0, current_idx - 2)
            self.term_start_combo.current(start_idx)
        else:
            self.term_end_combo.current(len(sorted_names) - 1)
            start_idx = max(0, len(sorted_names) - 3)
            self.term_start_combo.current(start_idx)

        self.status_var.set(
            f"{len(self.schools)} planteles, {len(sorted_names)} periodos de agosto cargados. "
            "Selecciona rango y pulsa 'Consultar Nuevo Ingreso'."
        )
        self._enable_controls()

    # --- Consulta ---
    def _on_query(self):
        start_sel = self.term_start_combo.current()
        end_sel = self.term_end_combo.current()
        if start_sel < 0 or end_sel < 0:
            messagebox.showwarning("Atención", "Selecciona periodo de inicio y de fin.")
            return
        if start_sel > end_sel:
            messagebox.showwarning("Atención",
                                   "El periodo de inicio debe ser anterior o igual al periodo de fin.")
            return

        self.selected_term_names = self._sorted_term_names[start_sel:end_sel + 1]
        num_terms = len(self.selected_term_names)
        num_schools = len(self.schools)
        total_queries = num_schools * num_terms

        self._cancel_event.clear()
        self._disable_controls()
        self.btn_cancel.configure(state="normal")
        self.btn_export.configure(state="disabled")
        self.progress["value"] = 0
        self.matrix = {}
        self.root.after(0, lambda: self._build_matrix_tree(self.selected_term_names))
        self.status_var.set(
            f"Consultando {num_schools} planteles × {num_terms} periodos ({total_queries} consultas)..."
        )

        def task():
            try:
                done = 0
                cancelled = False
                for i, school in enumerate(self.schools):
                    if self._cancel_event.is_set():
                        cancelled = True
                        break
                    school_id = school["id"]
                    school_name = school.get("name", f"School {school_id}")
                    school_cct = school.get("cct", "")
                    row_total = 0
                    school_terms = self._school_term_map.get(school_id, {})

                    for term_name in self.selected_term_names:
                        if self._cancel_event.is_set():
                            cancelled = True
                            break
                        self.root.after(0, lambda done=done, name=school_name, tname=term_name: (
                            self.status_var.set(
                                f"Plantel {name} — Periodo {tname}  ({done}/{total_queries})"
                            ),
                            self.progress.configure(value=(done / max(total_queries, 1)) * 100)
                        ))

                        term_id = school_terms.get(term_name)
                        if term_id:
                            try:
                                count = get_nuevo_ingreso_count(self.access_token, school_id, term_id)
                            except Exception:
                                count = 0
                        else:
                            count = 0

                        self.matrix[(school_id, term_name)] = count
                        row_total += count
                        done += 1

                    if cancelled:
                        break

                    values = [i + 1, school_cct, school_name]
                    for term_name in self.selected_term_names:
                        values.append(self.matrix.get((school_id, term_name), 0))
                    values.append(row_total)
                    self.root.after(0, lambda v=tuple(values): self.tree.insert("", tk.END, values=v))

                if cancelled:
                    self.root.after(0, lambda done=done: self._query_cancelled(done, total_queries))
                else:
                    self.root.after(0, self._query_done)
            except Exception as e:
                msg = str(e)
                self.root.after(0, lambda msg=msg: self._show_error("Error en consulta", msg))
            finally:
                self.root.after(0, self._enable_controls)

        threading.Thread(target=task, daemon=True).start()

    def _on_cancel(self):
        self._cancel_event.set()
        self.btn_cancel.configure(state="disabled")
        self.status_var.set("Cancelando...")

    def _query_cancelled(self, done, total):
        self.btn_cancel.configure(state="disabled")
        self.progress["value"] = (done / max(total, 1)) * 100
        self.status_var.set(f"Consulta cancelada: {done}/{total} consultas completadas.")
        if self.matrix:
            self.btn_export.configure(state="normal")

    def _query_done(self):
        self.btn_cancel.configure(state="disabled")
        self.progress["value"] = 100
        grand_total = sum(self.matrix.values())
        self.status_var.set(
            f"Consulta completada: {len(self.schools)} planteles × "
            f"{len(self.selected_term_names)} periodos. "
            f"Total nuevo ingreso: {grand_total} alumnos."
        )
        if self.matrix:
            self.btn_export.configure(state="normal")

    # --- Exportar a Excel ---
    def _on_export(self):
        if not self.matrix:
            messagebox.showwarning("Atención", "No hay datos para exportar. Realiza una consulta primero.")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Guardar Excel como...",
            initialfile="nuevo_ingreso_primer_semestre.xlsx"
        )
        if not filepath:
            return

        self.status_var.set("Exportando a Excel...")

        def task():
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from openpyxl.utils import get_column_letter

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Nuevo Ingreso"

                header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
                header_fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                bold_font = Font(name="Calibri", bold=True, size=11)
                center_align = Alignment(horizontal="center")
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

                term_names = self.selected_term_names
                num_terms = len(term_names)
                total_cols = 3 + num_terms + 1

                # Título (fila 1)
                t_start = term_names[0]
                t_end = term_names[-1]
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
                title_cell = ws.cell(
                    row=1, column=1,
                    value=f"Nuevo Ingreso — Alumnos de 1er Semestre — {t_start} a {t_end}"
                )
                title_cell.font = Font(name="Calibri", bold=True, size=14)
                title_cell.alignment = Alignment(horizontal="center")

                # Subtítulo (fila 2)
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
                sub_cell = ws.cell(
                    row=2, column=1,
                    value="Solo periodos que inician en agosto · grade_level = 1"
                )
                sub_cell.font = Font(name="Calibri", italic=True, size=10, color="555555")
                sub_cell.alignment = Alignment(horizontal="center")

                # Encabezados (fila 4)
                fixed_headers = ["#", "CCT", "Nombre del Plantel"]
                for col, h in enumerate(fixed_headers, 1):
                    cell = ws.cell(row=4, column=col, value=h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

                for j, name in enumerate(term_names):
                    col = 4 + j
                    cell = ws.cell(row=4, column=col, value=name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

                total_col = 4 + num_terms
                cell = ws.cell(row=4, column=total_col, value="TOTAL")
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

                # Datos (filas 5+)
                for i, school in enumerate(self.schools):
                    row = 5 + i
                    school_id = school["id"]
                    ws.cell(row=row, column=1, value=i + 1).border = thin_border
                    ws.cell(row=row, column=2, value=school.get("cct", "")).border = thin_border
                    ws.cell(row=row, column=3, value=school.get("name", "")).border = thin_border

                    row_total = 0
                    for j, term_name in enumerate(term_names):
                        count = self.matrix.get((school_id, term_name), 0)
                        row_total += count
                        c = ws.cell(row=row, column=4 + j, value=count)
                        c.border = thin_border
                        c.alignment = center_align

                    c = ws.cell(row=row, column=total_col, value=row_total)
                    c.border = thin_border
                    c.alignment = center_align
                    c.font = bold_font

                # Fila de totales por periodo
                totals_row = 5 + len(self.schools)
                ws.merge_cells(start_row=totals_row, start_column=1,
                               end_row=totals_row, end_column=3)
                lbl = ws.cell(row=totals_row, column=1, value="TOTAL POR PERIODO")
                lbl.font = bold_font
                lbl.alignment = Alignment(horizontal="right")
                lbl.border = thin_border
                for c in range(2, 4):
                    ws.cell(row=totals_row, column=c).border = thin_border

                grand_total = 0
                for j, term_name in enumerate(term_names):
                    col_total = sum(
                        self.matrix.get((s["id"], term_name), 0) for s in self.schools
                    )
                    grand_total += col_total
                    c = ws.cell(row=totals_row, column=4 + j, value=col_total)
                    c.border = thin_border
                    c.alignment = center_align
                    c.font = bold_font

                c = ws.cell(row=totals_row, column=total_col, value=grand_total)
                c.border = thin_border
                c.alignment = center_align
                c.font = Font(name="Calibri", bold=True, size=12)

                # Ancho de columnas
                ws.column_dimensions["A"].width = 6
                ws.column_dimensions["B"].width = 18
                ws.column_dimensions["C"].width = 45
                for j in range(num_terms):
                    letter = get_column_letter(4 + j)
                    ws.column_dimensions[letter].width = 16
                ws.column_dimensions[get_column_letter(total_col)].width = 12

                wb.save(filepath)

                self.root.after(0, lambda: self.status_var.set(f"Excel exportado: {filepath}"))
                self.root.after(0, lambda: messagebox.showinfo(
                    "Listo",
                    f"Excel exportado exitosamente.\n"
                    f"{len(self.schools)} planteles × {num_terms} periodos de agosto\n"
                    f"Total nuevo ingreso: {grand_total} alumnos\n"
                    f"Guardado en: {filepath}"
                ))
            except ImportError:
                self.root.after(0, lambda: messagebox.showerror(
                    "Módulo faltante",
                    "Se requiere 'openpyxl' para exportar a Excel.\n"
                    "Instálalo con: pip install openpyxl"
                ))
            except Exception as e:
                msg = str(e)
                self.root.after(0, lambda msg=msg: self._show_error("Error exportando Excel", msg))

        threading.Thread(target=task, daemon=True).start()

    # --- Helpers UI ---
    def _disable_controls(self):
        self.term_start_combo.configure(state="disabled")
        self.term_end_combo.configure(state="disabled")
        self.btn_query.configure(state="disabled")

    def _enable_controls(self):
        self.term_start_combo.configure(state="readonly")
        self.term_end_combo.configure(state="readonly")
        self.btn_query.configure(state="normal")

    def _show_error(self, title, message):
        self.status_var.set(f"Error: {message[:80]}")
        messagebox.showerror(title, message)
        self._enable_controls()


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    root = tk.Tk()
    app = SaekoNuevoIngresoApp(root)
    root.mainloop()
